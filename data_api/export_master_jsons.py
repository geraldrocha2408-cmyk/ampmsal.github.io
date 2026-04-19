from __future__ import annotations

import json
import math
import re
import sys
import traceback
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


VERSION_EXPORTER = "1.0.0"
SOURCE_FILENAME = "Cubo_Semanal_Compactado.xlsx"


# -----------------------------
# Logging
# -----------------------------
class Logger:
    @staticmethod
    def info(msg: str) -> None:
        print(f"[INFO] {msg}")

    @staticmethod
    def warn(msg: str) -> None:
        print(f"[WARN] {msg}")

    @staticmethod
    def error(msg: str) -> None:
        print(f"[ERROR] {msg}")


# -----------------------------
# Context / helpers
# -----------------------------
@dataclass
class ExportContext:
    project_root: Path
    source_file: Path
    out_root: Path
    warnings: List[str] = field(default_factory=list)
    files_generated: List[str] = field(default_factory=list)
    row_counts_by_file: Dict[str, int] = field(default_factory=dict)
    detected_sheets: List[str] = field(default_factory=list)
    months_available: List[str] = field(default_factory=list)
    weeks_available: List[str] = field(default_factory=list)
    stores_available: List[str] = field(default_factory=list)
    modules_skipped: Dict[str, str] = field(default_factory=dict)

    def warn(self, msg: str) -> None:
        self.warnings.append(msg)
        Logger.warn(msg)

    def add_file(self, rel_path: str, row_count: int) -> None:
        rel_path = rel_path.replace("\\", "/")
        self.files_generated.append(rel_path)
        self.row_counts_by_file[rel_path] = int(row_count)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def compact_key(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def safe_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return float(value)
    try:
        if isinstance(value, str):
            value = value.replace(",", "").strip()
        number = float(value)
        if math.isnan(number) or math.isinf(number):
            return default
        return number
    except Exception:
        return default


def safe_int(value: Any, default: int = 0) -> int:
    return int(round(safe_float(value, default)))


def safe_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = normalize_text(value)
    return text in {"1", "true", "si", "yes", "y", "verdadero", "x"}


def nonempty_str(value: Any, default: Optional[str] = None) -> Optional[str]:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def safe_div(num: float, den: float) -> float:
    if den in (0, None):
        return 0.0
    try:
        return num / den
    except Exception:
        return 0.0


def round_value(value: Any) -> Any:
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return 0.0
        return round(value, 6)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def json_ready_records(records: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for rec in records:
        clean = {k: round_value(v) for k, v in rec.items()}
        out.append(clean)
    return out


def ensure_dirs(base: Path) -> None:
    folders = [
        "meta",
        "ventas",
        "delivery",
        "inventario",
        "merma",
        "sbf",
        "cxc",
        "prd0",
        "hallazgos",
        "innovation",
    ]
    for folder in folders:
        (base / folder).mkdir(parents=True, exist_ok=True)


def write_json(ctx: ExportContext, relative_path: str, payload: Any) -> None:
    path = ctx.out_root / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    row_count = len(payload) if isinstance(payload, list) else len(payload) if isinstance(payload, dict) else 1
    ctx.add_file(path.relative_to(ctx.project_root).as_posix(), row_count)
    Logger.info(f"JSON exportado: {path.relative_to(ctx.project_root).as_posix()} ({row_count} registros)")


# -----------------------------
# Date / period helpers
# -----------------------------
MONTH_RE = re.compile(r"^(\d{4})-(\d{2})$")
WEEK_RE = re.compile(r"^(\d{4})-W(\d{2})$")


def parse_month_key(month_key: str) -> Optional[date]:
    if not month_key:
        return None
    m = MONTH_RE.match(str(month_key).strip())
    if not m:
        return None
    return date(int(m.group(1)), int(m.group(2)), 1)


def add_months(month_key: str, months: int) -> Optional[str]:
    dt = parse_month_key(month_key)
    if not dt:
        return None
    idx = (dt.year * 12 + dt.month - 1) + months
    year = idx // 12
    month = idx % 12 + 1
    return f"{year:04d}-{month:02d}"


def prev_month_key(month_key: str) -> Optional[str]:
    return add_months(month_key, -1)


def ly_month_key(month_key: str) -> Optional[str]:
    return add_months(month_key, -12)


def parse_week_key(week_key: str) -> Optional[Tuple[int, int]]:
    if not week_key:
        return None
    m = WEEK_RE.match(str(week_key).strip())
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def iso_week_to_monday(week_key: str) -> Optional[date]:
    parsed = parse_week_key(week_key)
    if not parsed:
        return None
    year, week = parsed
    try:
        return date.fromisocalendar(year, week, 1)
    except Exception:
        return None


def week_key_from_date(dt: date) -> str:
    iso = dt.isocalendar()
    return f"{iso.year:04d}-W{iso.week:02d}"


def shift_week_key(week_key: str, weeks_delta: int) -> Optional[str]:
    monday = iso_week_to_monday(week_key)
    if not monday:
        return None
    return week_key_from_date(monday + timedelta(weeks=weeks_delta))


def ly_week_key(week_key: str) -> Optional[str]:
    return shift_week_key(week_key, -52)


def make_dow_group(dow_name: Optional[str], dow_number: Optional[int]) -> Optional[str]:
    if dow_number is not None and dow_number != 0:
        return "Lun-Jue" if int(dow_number) <= 4 else "Vie-Dom"
    norm = normalize_text(dow_name)
    if norm in {"lunes", "martes", "miercoles", "jueves"}:
        return "Lun-Jue"
    if norm in {"viernes", "sabado", "domingo"}:
        return "Vie-Dom"
    return None


# -----------------------------
# Workbook reader
# -----------------------------
class WorkbookReader:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self.wb = load_workbook(workbook_path, read_only=True, data_only=True)
        self.sheetnames = list(self.wb.sheetnames)
        self.sheet_key_map = {compact_key(name): name for name in self.sheetnames}
        self._headers_cache: Dict[str, List[Any]] = {}

    def find_sheet(self, *candidates: str) -> Optional[str]:
        keys = {compact_key(c) for c in candidates if c}
        for key, actual in self.sheet_key_map.items():
            if key in keys:
                return actual
        # contains fallback
        for candidate in candidates:
            ckey = compact_key(candidate)
            for key, actual in self.sheet_key_map.items():
                if ckey and (ckey in key or key in ckey):
                    return actual
        return None

    def has_sheet(self, *candidates: str) -> bool:
        return self.find_sheet(*candidates) is not None

    def get_headers(self, sheet_name: str) -> List[Any]:
        if sheet_name not in self._headers_cache:
            ws = self.wb[sheet_name]
            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            self._headers_cache[sheet_name] = list(first_row)
        return self._headers_cache[sheet_name]

    def resolve_columns(
        self,
        sheet_name: str,
        alias_map: Dict[str, Sequence[str]],
    ) -> Dict[str, int]:
        headers = self.get_headers(sheet_name)
        normalized_headers = [normalize_text(h) for h in headers]
        col_map: Dict[str, int] = {}
        for canonical, aliases in alias_map.items():
            found_idx: Optional[int] = None
            alias_norms = [normalize_text(a) for a in aliases]
            # exact match first
            for idx, hnorm in enumerate(normalized_headers):
                if hnorm in alias_norms:
                    found_idx = idx
                    break
            # contains fallback
            if found_idx is None:
                for idx, hnorm in enumerate(normalized_headers):
                    if any(alias and (alias in hnorm or hnorm in alias) for alias in alias_norms):
                        found_idx = idx
                        break
            if found_idx is not None:
                col_map[canonical] = found_idx
        return col_map

    def iter_rows(self, sheet_name: str, col_map: Dict[str, int]) -> Iterable[Dict[str, Any]]:
        ws = self.wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            yield {canonical: row[idx] if idx < len(row) else None for canonical, idx in col_map.items()}


# -----------------------------
# Aggregation utilities
# -----------------------------
def metric_bucket_factory() -> Dict[str, float]:
    return defaultdict(float)  # type: ignore[return-value]


def add_metric(bucket: Dict[str, float], key: str, value: Any) -> None:
    bucket[key] = bucket.get(key, 0.0) + safe_float(value)


def finalize_metric_dict(d: Dict[Any, Dict[str, Any]], key_fields: Sequence[str]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for key, metrics in d.items():
        if not isinstance(key, tuple):
            key = (key,)
        row = {field: value for field, value in zip(key_fields, key)}
        row.update(metrics)
        rows.append(row)
    return rows


def apply_period_comparison(
    records: List[Dict[str, Any]],
    period_field: str,
    grouping_fields: Sequence[str],
    metric_fields: Sequence[str],
    period_type: str,
) -> List[Dict[str, Any]]:
    idx: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for rec in records:
        key = tuple(rec.get(f) for f in (*grouping_fields, period_field))
        idx[key] = rec

    for rec in records:
        period_key = rec.get(period_field)
        if period_type == "month":
            compare_period = ly_month_key(str(period_key)) if period_key else None
            prev_period = prev_month_key(str(period_key)) if period_key else None
        else:
            compare_period = ly_week_key(str(period_key)) if period_key else None
            prev_period = shift_week_key(str(period_key), -1) if period_key else None

        compare_key = tuple(rec.get(f) for f in grouping_fields) + (compare_period,)
        prev_key = tuple(rec.get(f) for f in grouping_fields) + (prev_period,)
        compare_rec = idx.get(compare_key, {}) if compare_period else {}
        prev_rec = idx.get(prev_key, {}) if prev_period else {}

        for metric in metric_fields:
            current_value = safe_float(rec.get(metric))
            compare_value = safe_float(compare_rec.get(metric))
            prev_value = safe_float(prev_rec.get(metric))
            rec[f"{metric}_ly"] = compare_value
            rec[f"{metric}_delta"] = current_value - compare_value
            rec[f"{metric}_delta_pct"] = safe_div(current_value - compare_value, compare_value)
            rec[f"{metric}_lm"] = prev_value
            rec[f"{metric}_vs_lm"] = current_value - prev_value
            rec[f"{metric}_vs_lm_pct"] = safe_div(current_value - prev_value, prev_value)
    return records


def sort_records(records: List[Dict[str, Any]], keys: Sequence[str]) -> List[Dict[str, Any]]:
    return sorted(records, key=lambda x: tuple("" if x.get(k) is None else x.get(k) for k in keys))


def enrich_sales_common(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    for rec in records:
        sales = safe_float(rec.get("sales"))
        margin = safe_float(rec.get("margin"))
        tx = safe_float(rec.get("tx"))
        rec["margin_pct"] = safe_div(margin, sales)
        rec["avg_ticket"] = safe_div(sales, tx)
        sales_ly = safe_float(rec.get("sales_ly"))
        margin_ly = safe_float(rec.get("margin_ly"))
        tx_ly = safe_float(rec.get("tx_ly"))
        rec["margin_pct_ly"] = safe_div(margin_ly, sales_ly)
        rec["avg_ticket_ly"] = safe_div(sales_ly, tx_ly)
    return records


def severity_from_pct(delta_pct: float, reverse: bool = False) -> str:
    value = -delta_pct if reverse else delta_pct
    abs_value = abs(delta_pct)
    if value <= -0.20 or abs_value >= 0.35:
        return "high"
    if value <= -0.10 or abs_value >= 0.20:
        return "medium"
    return "low"


# -----------------------------
# Domain processors
# -----------------------------
def process_sales_and_delivery(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando ventas y delivery...")
    sales_sheet = reader.find_sheet("Cubo_Sales_DetailHallazgos")
    if not sales_sheet:
        ctx.warn("No se encontró hoja de ventas detalle. Se omiten exportaciones de ventas y delivery basadas en ventas.")
        return {}

    alias_map = {
        "date_sold": ["date sold", "date", "fecha"],
        "week_key": ["yearweek", "week", "year month"],
        "month_key": ["calmonth", "month", "mes", "calendar month"],
        "store": ["store", "tienda", "store name"],
        "department": ["department", "departamento"],
        "category": ["category", "categoria", "categoría"],
        "supplier": ["supplier", "proveedor"],
        "brand": ["brand", "marca"],
        "description": ["description", "descripcion", "descripción", "descripcion3"],
        "daypart": ["daypart", "part of day", "turno"],
        "dow": ["dow", "dayofweek", "dia", "día"],
        "dow_name": ["dow name", "dow_name", "weekday", "nombre dia", "dia nombre", "day name"],
        "is_delivery": ["is delivery", "delivery", "es delivery"],
        "platform": ["delivery channel", "platform", "canal", "delivery platform"],
        "qty": ["qty sold", "qty", "quantity", "qty ven", "unidades"],
        "sales": ["sales", "total sales", "venta", "ventas"],
        "margin": ["total gross margin", "gross margin", "margin", "margen"],
    }
    col_map = reader.resolve_columns(sales_sheet, alias_map)
    critical = ["week_key", "month_key", "store", "department", "category", "description", "sales", "margin", "qty"]
    missing = [c for c in critical if c not in col_map]
    if missing:
        ctx.warn(f"Hoja {sales_sheet}: faltan columnas críticas para ventas/delivery: {missing}. Se omite módulo.")
        return {}

    monthly_business: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    weekly_business: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    monthly_store: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    weekly_store: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    monthly_department: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    weekly_department: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    monthly_category: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    weekly_category: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    monthly_brand: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    weekly_brand: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    monthly_description: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    weekly_description: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    sales_by_dow: Dict[Any, Dict[str, Any]] = defaultdict(dict)

    delivery_monthly_business: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    delivery_weekly_business: Dict[Any, Dict[str, Any]] = defaultdict(metric_bucket_factory)
    delivery_monthly_store_platform: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    delivery_weekly_store_platform: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    delivery_monthly_category: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    delivery_weekly_category: Dict[Any, Dict[str, Any]] = defaultdict(dict)
    delivery_by_dow: Dict[Any, Dict[str, Any]] = defaultdict(dict)

    stores, months, weeks = set(), set(), set()
    departments, categories, brands, descriptions, platforms, suppliers = set(), set(), set(), set(), set(), set()

    def seed_dim_bucket(container: Dict[Any, Dict[str, Any]], key: Tuple[Any, ...], extra: Dict[str, Any]) -> Dict[str, Any]:
        if key not in container or not container[key]:
            container[key] = {**extra, "sales": 0.0, "margin": 0.0, "qty": 0.0}
        return container[key]

    row_count = 0
    for row in reader.iter_rows(sales_sheet, col_map):
        row_count += 1
        month_key = nonempty_str(row.get("month_key"))
        week_key = nonempty_str(row.get("week_key"))
        store = nonempty_str(row.get("store"), "UNKNOWN")
        department = nonempty_str(row.get("department"), "Sin departamento")
        category = nonempty_str(row.get("category"), "Sin categoría")
        brand = nonempty_str(row.get("brand"), "Sin marca")
        description = nonempty_str(row.get("description"), "Sin descripción")
        supplier = nonempty_str(row.get("supplier"), "Sin proveedor")
        platform = nonempty_str(row.get("platform"), "No Delivery") if safe_bool(row.get("is_delivery")) or row.get("platform") else None
        is_delivery = safe_bool(row.get("is_delivery")) or bool(platform)
        dow = safe_int(row.get("dow"), 0)
        dow_name = nonempty_str(row.get("dow_name"))
        dow_group = make_dow_group(dow_name, dow)
        sales = safe_float(row.get("sales"))
        margin = safe_float(row.get("margin"))
        qty = safe_float(row.get("qty"))

        if month_key:
            months.add(month_key)
        if week_key:
            weeks.add(week_key)
        stores.add(store)
        departments.add(department)
        categories.add(category)
        brands.add(brand)
        descriptions.add(description)
        suppliers.add(supplier)
        if platform:
            platforms.add(platform)

        if month_key:
            add_metric(monthly_business[month_key], "sales", sales)
            add_metric(monthly_business[month_key], "margin", margin)
            add_metric(monthly_business[month_key], "qty", qty)
            add_metric(monthly_store[(month_key, store)], "sales", sales)
            add_metric(monthly_store[(month_key, store)], "margin", margin)
            add_metric(monthly_store[(month_key, store)], "qty", qty)
            add_metric(monthly_department[(month_key, department)], "sales", sales)
            add_metric(monthly_department[(month_key, department)], "margin", margin)
            add_metric(monthly_department[(month_key, department)], "qty", qty)
            add_metric(monthly_category[(month_key, department, category)], "sales", sales)
            add_metric(monthly_category[(month_key, department, category)], "margin", margin)
            add_metric(monthly_category[(month_key, department, category)], "qty", qty)
            add_metric(monthly_brand[(month_key, department, category, brand)], "sales", sales)
            add_metric(monthly_brand[(month_key, department, category, brand)], "margin", margin)
            add_metric(monthly_brand[(month_key, department, category, brand)], "qty", qty)
            mb = seed_dim_bucket(
                monthly_description,
                (month_key, store, description),
                {
                    "month_key": month_key,
                    "store": store,
                    "store_name": store,
                    "department": department,
                    "category": category,
                    "brand": brand,
                    "description": description,
                    "supplier": supplier,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                },
            )
            mb["sales"] += sales
            mb["margin"] += margin
            mb["qty"] += qty

        if week_key:
            add_metric(weekly_business[week_key], "sales", sales)
            add_metric(weekly_business[week_key], "margin", margin)
            add_metric(weekly_business[week_key], "qty", qty)
            add_metric(weekly_store[(week_key, store)], "sales", sales)
            add_metric(weekly_store[(week_key, store)], "margin", margin)
            add_metric(weekly_store[(week_key, store)], "qty", qty)
            add_metric(weekly_department[(week_key, department)], "sales", sales)
            add_metric(weekly_department[(week_key, department)], "margin", margin)
            add_metric(weekly_department[(week_key, department)], "qty", qty)
            add_metric(weekly_category[(week_key, department, category)], "sales", sales)
            add_metric(weekly_category[(week_key, department, category)], "margin", margin)
            add_metric(weekly_category[(week_key, department, category)], "qty", qty)
            add_metric(weekly_brand[(week_key, department, category, brand)], "sales", sales)
            add_metric(weekly_brand[(week_key, department, category, brand)], "margin", margin)
            add_metric(weekly_brand[(week_key, department, category, brand)], "qty", qty)
            wb_desc = seed_dim_bucket(
                weekly_description,
                (week_key, store, description),
                {
                    "week_key": week_key,
                    "store": store,
                    "store_name": store,
                    "department": department,
                    "category": category,
                    "brand": brand,
                    "description": description,
                    "supplier": supplier,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                },
            )
            wb_desc["sales"] += sales
            wb_desc["margin"] += margin
            wb_desc["qty"] += qty

        if month_key:
            key = ("month", month_key, store, dow, dow_name or f"DOW {dow}")
            bucket = sales_by_dow.get(key) or {
                "period_type": "month",
                "period_key": month_key,
                "store": store,
                "store_name": store,
                "dow": dow,
                "dow_name": dow_name,
                "dow_group": dow_group,
                "sales": 0.0,
                "margin": 0.0,
                "qty": 0.0,
            }
            bucket["sales"] += sales
            bucket["margin"] += margin
            bucket["qty"] += qty
            sales_by_dow[key] = bucket
        if week_key:
            key = ("week", week_key, store, dow, dow_name or f"DOW {dow}")
            bucket = sales_by_dow.get(key) or {
                "period_type": "week",
                "period_key": week_key,
                "store": store,
                "store_name": store,
                "dow": dow,
                "dow_name": dow_name,
                "dow_group": dow_group,
                "sales": 0.0,
                "margin": 0.0,
                "qty": 0.0,
            }
            bucket["sales"] += sales
            bucket["margin"] += margin
            bucket["qty"] += qty
            sales_by_dow[key] = bucket

        if is_delivery and platform:
            if month_key:
                add_metric(delivery_monthly_business[month_key], "sales", sales)
                add_metric(delivery_monthly_business[month_key], "margin", margin)
                add_metric(delivery_monthly_business[month_key], "qty", qty)
                msp = delivery_monthly_store_platform.get((month_key, store, platform)) or {
                    "month_key": month_key,
                    "store": store,
                    "store_name": store,
                    "platform": platform,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                msp["sales"] += sales
                msp["margin"] += margin
                msp["qty"] += qty
                delivery_monthly_store_platform[(month_key, store, platform)] = msp
                mcat = delivery_monthly_category.get((month_key, platform, category)) or {
                    "month_key": month_key,
                    "platform": platform,
                    "category": category,
                    "department": department,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                mcat["sales"] += sales
                mcat["margin"] += margin
                mcat["qty"] += qty
                delivery_monthly_category[(month_key, platform, category)] = mcat
                mdow = delivery_by_dow.get(("month", month_key, store, platform, dow)) or {
                    "period_type": "month",
                    "period_key": month_key,
                    "store": store,
                    "store_name": store,
                    "platform": platform,
                    "dow": dow,
                    "dow_name": dow_name,
                    "dow_group": dow_group,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                mdow["sales"] += sales
                mdow["margin"] += margin
                mdow["qty"] += qty
                delivery_by_dow[("month", month_key, store, platform, dow)] = mdow

            if week_key:
                add_metric(delivery_weekly_business[week_key], "sales", sales)
                add_metric(delivery_weekly_business[week_key], "margin", margin)
                add_metric(delivery_weekly_business[week_key], "qty", qty)
                wsp = delivery_weekly_store_platform.get((week_key, store, platform)) or {
                    "week_key": week_key,
                    "store": store,
                    "store_name": store,
                    "platform": platform,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                wsp["sales"] += sales
                wsp["margin"] += margin
                wsp["qty"] += qty
                delivery_weekly_store_platform[(week_key, store, platform)] = wsp
                wcat = delivery_weekly_category.get((week_key, platform, category)) or {
                    "week_key": week_key,
                    "platform": platform,
                    "category": category,
                    "department": department,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                wcat["sales"] += sales
                wcat["margin"] += margin
                wcat["qty"] += qty
                delivery_weekly_category[(week_key, platform, category)] = wcat
                wdow = delivery_by_dow.get(("week", week_key, store, platform, dow)) or {
                    "period_type": "week",
                    "period_key": week_key,
                    "store": store,
                    "store_name": store,
                    "platform": platform,
                    "dow": dow,
                    "dow_name": dow_name,
                    "dow_group": dow_group,
                    "sales": 0.0,
                    "margin": 0.0,
                    "qty": 0.0,
                }
                wdow["sales"] += sales
                wdow["margin"] += margin
                wdow["qty"] += qty
                delivery_by_dow[("week", week_key, store, platform, dow)] = wdow

    Logger.info(f"Ventas detalle procesadas: {row_count} filas")

    sales_outputs = {
        "monthly_business": sort_records(
            apply_period_comparison(
                [{"month_key": k, **v} for k, v in monthly_business.items()],
                "month_key",
                [],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key"],
        ),
        "weekly_business": sort_records(
            apply_period_comparison(
                [{"week_key": k, **v} for k, v in weekly_business.items()],
                "week_key",
                [],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key"],
        ),
        "monthly_store": sort_records(
            apply_period_comparison(
                [{"month_key": k[0], "store": k[1], "store_name": k[1], **v} for k, v in monthly_store.items()],
                "month_key",
                ["store"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "store"],
        ),
        "weekly_store": sort_records(
            apply_period_comparison(
                [{"week_key": k[0], "store": k[1], "store_name": k[1], **v} for k, v in weekly_store.items()],
                "week_key",
                ["store"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "store"],
        ),
        "monthly_department": sort_records(
            apply_period_comparison(
                [{"month_key": k[0], "department": k[1], **v} for k, v in monthly_department.items()],
                "month_key",
                ["department"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "department"],
        ),
        "weekly_department": sort_records(
            apply_period_comparison(
                [{"week_key": k[0], "department": k[1], **v} for k, v in weekly_department.items()],
                "week_key",
                ["department"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "department"],
        ),
        "monthly_category": sort_records(
            apply_period_comparison(
                [{"month_key": k[0], "department": k[1], "category": k[2], **v} for k, v in monthly_category.items()],
                "month_key",
                ["department", "category"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "department", "category"],
        ),
        "weekly_category": sort_records(
            apply_period_comparison(
                [{"week_key": k[0], "department": k[1], "category": k[2], **v} for k, v in weekly_category.items()],
                "week_key",
                ["department", "category"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "department", "category"],
        ),
        "monthly_brand": sort_records(
            apply_period_comparison(
                [{"month_key": k[0], "department": k[1], "category": k[2], "brand": k[3], **v} for k, v in monthly_brand.items()],
                "month_key",
                ["department", "category", "brand"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "department", "category", "brand"],
        ),
        "weekly_brand": sort_records(
            apply_period_comparison(
                [{"week_key": k[0], "department": k[1], "category": k[2], "brand": k[3], **v} for k, v in weekly_brand.items()],
                "week_key",
                ["department", "category", "brand"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "department", "category", "brand"],
        ),
        "monthly_description": sort_records(
            apply_period_comparison(list(monthly_description.values()), "month_key", ["store", "description"], ["sales", "margin", "qty"], "month"),
            ["month_key", "store", "description"],
        ),
        "weekly_description": sort_records(
            apply_period_comparison(list(weekly_description.values()), "week_key", ["store", "description"], ["sales", "margin", "qty"], "week"),
            ["week_key", "store", "description"],
        ),
        "sales_by_dow": sort_records(list(sales_by_dow.values()), ["period_type", "period_key", "store", "dow"]),
    }

    delivery_outputs = {
        "monthly_business": sort_records(
            apply_period_comparison(
                [{"month_key": k, **v} for k, v in delivery_monthly_business.items()],
                "month_key",
                [],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key"],
        ),
        "weekly_business": sort_records(
            apply_period_comparison(
                [{"week_key": k, **v} for k, v in delivery_weekly_business.items()],
                "week_key",
                [],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key"],
        ),
        "monthly_store_platform": sort_records(
            apply_period_comparison(
                list(delivery_monthly_store_platform.values()),
                "month_key",
                ["store", "platform"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "store", "platform"],
        ),
        "weekly_store_platform": sort_records(
            apply_period_comparison(
                list(delivery_weekly_store_platform.values()),
                "week_key",
                ["store", "platform"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "store", "platform"],
        ),
        "monthly_category": sort_records(
            apply_period_comparison(
                list(delivery_monthly_category.values()),
                "month_key",
                ["platform", "category"],
                ["sales", "margin", "qty"],
                "month",
            ),
            ["month_key", "platform", "category"],
        ),
        "weekly_category": sort_records(
            apply_period_comparison(
                list(delivery_weekly_category.values()),
                "week_key",
                ["platform", "category"],
                ["sales", "margin", "qty"],
                "week",
            ),
            ["week_key", "platform", "category"],
        ),
        "delivery_by_dow": sort_records(list(delivery_by_dow.values()), ["period_type", "period_key", "store", "platform", "dow"]),
    }

    return {
        "sales_outputs": sales_outputs,
        "delivery_outputs": delivery_outputs,
        "catalogs": {
            "stores": sorted(stores),
            "months": sorted(months),
            "weeks": sorted(weeks),
            "departments": sorted(departments),
            "categories": sorted(categories),
            "brands": sorted(brands),
            "descriptions": sorted(descriptions),
            "delivery_platforms": sorted(platforms),
            "suppliers": sorted(suppliers),
        },
    }


def merge_tx_into_sales_outputs(ctx: ExportContext, reader: WorkbookReader, results: Dict[str, Any]) -> None:
    Logger.info("Procesando TX para enriquecer ventas y delivery...")
    weekly_sheet = reader.find_sheet("Cubo_TX_Fact")
    monthly_sheet = reader.find_sheet("Cubo_TX_Month_Fact")
    if not weekly_sheet and not monthly_sheet:
        ctx.warn("No se encontraron hojas TX. Los JSON de ventas/delivery quedarán sin transacciones.")
        return

    weekly_aliases = {
        "week_key": ["yearweek", "yearmonth", "week"],
        "store": ["store", "tienda"],
        "is_delivery": ["is delivery", "delivery"],
        "platform": ["delivery channel", "platform", "canal"],
        "tx": ["transactions", "tx", "transacciones"],
    }
    monthly_aliases = {
        "month_key": ["calmonth", "yearmonth", "month"],
        "store": ["store", "tienda"],
        "is_delivery": ["is delivery", "delivery"],
        "platform": ["delivery channel", "platform", "canal"],
        "tx": ["transactions", "tx", "transacciones"],
    }

    weekly_business_tx: Dict[str, float] = defaultdict(float)
    weekly_store_tx: Dict[Tuple[str, str], float] = defaultdict(float)
    weekly_delivery_business_tx: Dict[str, float] = defaultdict(float)
    weekly_delivery_store_platform_tx: Dict[Tuple[str, str, str], float] = defaultdict(float)

    monthly_business_tx: Dict[str, float] = defaultdict(float)
    monthly_store_tx: Dict[Tuple[str, str], float] = defaultdict(float)
    monthly_delivery_business_tx: Dict[str, float] = defaultdict(float)
    monthly_delivery_store_platform_tx: Dict[Tuple[str, str, str], float] = defaultdict(float)

    if weekly_sheet:
        col_map = reader.resolve_columns(weekly_sheet, weekly_aliases)
        if all(k in col_map for k in ["week_key", "store", "tx"]):
            for row in reader.iter_rows(weekly_sheet, col_map):
                week_key = nonempty_str(row.get("week_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                platform = nonempty_str(row.get("platform")) if row.get("platform") else None
                is_delivery = safe_bool(row.get("is_delivery")) or bool(platform)
                tx = safe_float(row.get("tx"))
                if week_key:
                    weekly_business_tx[week_key] += tx
                    weekly_store_tx[(week_key, store)] += tx
                    if is_delivery and platform:
                        weekly_delivery_business_tx[week_key] += tx
                        weekly_delivery_store_platform_tx[(week_key, store, platform)] += tx
        else:
            ctx.warn(f"Hoja {weekly_sheet}: faltan columnas para TX semanal.")

    if monthly_sheet:
        col_map = reader.resolve_columns(monthly_sheet, monthly_aliases)
        if all(k in col_map for k in ["month_key", "store", "tx"]):
            for row in reader.iter_rows(monthly_sheet, col_map):
                month_key = nonempty_str(row.get("month_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                platform = nonempty_str(row.get("platform")) if row.get("platform") else None
                is_delivery = safe_bool(row.get("is_delivery")) or bool(platform)
                tx = safe_float(row.get("tx"))
                if month_key:
                    monthly_business_tx[month_key] += tx
                    monthly_store_tx[(month_key, store)] += tx
                    if is_delivery and platform:
                        monthly_delivery_business_tx[month_key] += tx
                        monthly_delivery_store_platform_tx[(month_key, store, platform)] += tx
        else:
            ctx.warn(f"Hoja {monthly_sheet}: faltan columnas para TX mensual.")

    def merge_tx(records: List[Dict[str, Any]], period_field: str, tx_map: Dict[Any, float], key_fields: Sequence[str]) -> List[Dict[str, Any]]:
        for rec in records:
            key = tuple(rec.get(f) for f in [period_field, *key_fields])
            if len(key) == 1:
                key = key[0]
            rec["tx"] = safe_float(tx_map.get(key, 0.0))
        return records

    sales_outputs = results.get("sales_outputs", {})
    delivery_outputs = results.get("delivery_outputs", {})

    sales_outputs["monthly_business"] = enrich_sales_common(merge_tx(sales_outputs.get("monthly_business", []), "month_key", monthly_business_tx, []))
    sales_outputs["weekly_business"] = enrich_sales_common(merge_tx(sales_outputs.get("weekly_business", []), "week_key", weekly_business_tx, []))
    sales_outputs["monthly_store"] = enrich_sales_common(merge_tx(sales_outputs.get("monthly_store", []), "month_key", monthly_store_tx, ["store"]))
    sales_outputs["weekly_store"] = enrich_sales_common(merge_tx(sales_outputs.get("weekly_store", []), "week_key", weekly_store_tx, ["store"]))

    delivery_outputs["monthly_business"] = enrich_sales_common(merge_tx(delivery_outputs.get("monthly_business", []), "month_key", monthly_delivery_business_tx, []))
    delivery_outputs["weekly_business"] = enrich_sales_common(merge_tx(delivery_outputs.get("weekly_business", []), "week_key", weekly_delivery_business_tx, []))
    delivery_outputs["monthly_store_platform"] = enrich_sales_common(merge_tx(delivery_outputs.get("monthly_store_platform", []), "month_key", monthly_delivery_store_platform_tx, ["store", "platform"]))
    delivery_outputs["weekly_store_platform"] = enrich_sales_common(merge_tx(delivery_outputs.get("weekly_store_platform", []), "week_key", weekly_delivery_store_platform_tx, ["store", "platform"]))

    # Add tx deltas after tx merge for required files
    for key, period_field, group_fields, period_type in [
        ("monthly_business", "month_key", [], "month"),
        ("weekly_business", "week_key", [], "week"),
        ("monthly_store", "month_key", ["store"], "month"),
        ("weekly_store", "week_key", ["store"], "week"),
    ]:
        if key in sales_outputs:
            apply_period_comparison(sales_outputs[key], period_field, group_fields, ["tx"], period_type)
            enrich_sales_common(sales_outputs[key])
    for key, period_field, group_fields, period_type in [
        ("monthly_business", "month_key", [], "month"),
        ("weekly_business", "week_key", [], "week"),
        ("monthly_store_platform", "month_key", ["store", "platform"], "month"),
        ("weekly_store_platform", "week_key", ["store", "platform"], "week"),
    ]:
        if key in delivery_outputs:
            apply_period_comparison(delivery_outputs[key], period_field, group_fields, ["tx"], period_type)
            enrich_sales_common(delivery_outputs[key])


def process_inventory(ctx: ExportContext, reader: WorkbookReader, results: Dict[str, Any]) -> Dict[str, Any]:
    Logger.info("Procesando inventario...")
    sheet = reader.find_sheet("Inventario_Detail")
    if not sheet:
        ctx.warn("No se encontró hoja Inventario_Detail. Se omite módulo inventario.")
        return {}

    alias_map = {
        "store": ["store", "tienda"],
        "department": ["department", "departamento"],
        "supplier": ["supplier", "proveedor"],
        "category": ["category", "categoria", "categoría"],
        "brand": ["brand", "marca"],
        "description": ["description", "descripcion", "descripción"],
        "cogs": ["cogs", "costo"],
        "amount": ["amount", "monto", "inventario $"],
        "quantity": ["quantity", "qty", "stock units", "cantidad"],
        "doi": ["dias inv", "días inv", "doi", "dias inventario"],
        "planimetria": ["planimetria", "planimetría"],
        "qty_sold_s4": ["qty sold s4", "qty s4"],
        "sales_s4": ["sales s4", "venta s4"],
        "unit_cost": ["costo unitario inv", "unit cost"],
        "ideal_qty": ["idealqty", "ideal qty"],
        "target_qty": ["targetqty", "target qty"],
        "ideal_amount": ["inv ideal $", "inv ideal", "ideal amount"],
        "ideal_doi": ["dias inv ideal", "días inv ideal", "ideal doi"],
    }
    col_map = reader.resolve_columns(sheet, alias_map)
    critical = ["store", "department", "supplier", "category", "brand", "description", "amount", "quantity", "doi"]
    missing = [c for c in critical if c not in col_map]
    if missing:
        ctx.warn(f"Hoja {sheet}: faltan columnas críticas para inventario: {missing}. Se omite módulo.")
        return {}

    snapshot: List[Dict[str, Any]] = []
    doi_rows: List[Dict[str, Any]] = []
    shortage_rows: List[Dict[str, Any]] = []
    excess_rows: List[Dict[str, Any]] = []
    transfer_pool: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for row in reader.iter_rows(sheet, col_map):
        store = nonempty_str(row.get("store"), "UNKNOWN")
        department = nonempty_str(row.get("department"), "Sin departamento")
        supplier = nonempty_str(row.get("supplier"), "Sin proveedor")
        category = nonempty_str(row.get("category"), "Sin categoría")
        brand = nonempty_str(row.get("brand"), "Sin marca")
        description = nonempty_str(row.get("description"), "Sin descripción")
        stock_units = safe_float(row.get("quantity"))
        inventory_amount = safe_float(row.get("amount"))
        cogs = safe_float(row.get("cogs"))
        doi = safe_float(row.get("doi"))
        planimetria = safe_float(row.get("planimetria"))
        qty_sold_s4 = safe_float(row.get("qty_sold_s4"))
        sales_s4 = safe_float(row.get("sales_s4"))
        unit_cost = safe_float(row.get("unit_cost"))
        ideal_qty = safe_float(row.get("ideal_qty"))
        target_qty = safe_float(row.get("target_qty")) or ideal_qty or planimetria
        ideal_amount = safe_float(row.get("ideal_amount"))
        ideal_doi = safe_float(row.get("ideal_doi"))
        avg_daily_qty = safe_div(qty_sold_s4, 28.0)
        shortage_units = max(target_qty - stock_units, 0.0)
        excess_units = max(stock_units - target_qty, 0.0)

        base = {
            "store": store,
            "store_name": store,
            "department": department,
            "category": category,
            "brand": brand,
            "description": description,
            "supplier": supplier,
            "stock_units": stock_units,
            "qty": stock_units,
            "inventory_amount": inventory_amount,
            "cogs": cogs,
            "doi": doi,
            "planimetria": planimetria,
            "qty_sold_s4": qty_sold_s4,
            "sales_s4": sales_s4,
            "avg_daily_qty_s4": avg_daily_qty,
            "unit_cost": unit_cost,
            "ideal_qty": ideal_qty,
            "target_qty": target_qty,
            "inventory_ideal_amount": ideal_amount,
            "doi_ideal": ideal_doi,
        }
        snapshot.append(base)
        doi_rows.append({**base, "priority": "inventory", "severity": "high" if doi >= 28 or doi <= 7 else "medium" if doi >= 21 or doi <= 10 else "low"})

        if (stock_units <= 0 and avg_daily_qty > 0) or (doi > 0 and doi < 7) or shortage_units >= max(1.0, target_qty * 0.35):
            severity = "high" if stock_units <= 0 or doi < 4 or shortage_units >= max(2.0, target_qty * 0.6) else "medium"
            shortage_rows.append(
                {
                    **base,
                    "shortage_units": shortage_units,
                    "share": safe_div(shortage_units, target_qty),
                    "priority": "shortage",
                    "severity": severity,
                }
            )

        if doi >= 28 or excess_units >= max(1.0, target_qty * 0.5):
            severity = "high" if doi >= 42 or excess_units >= max(3.0, target_qty) else "medium"
            excess_rows.append(
                {
                    **base,
                    "excess_units": excess_units,
                    "share": safe_div(excess_units, target_qty if target_qty else stock_units or 1),
                    "priority": "excess",
                    "severity": severity,
                }
            )

        transfer_pool[description].append(
            {
                **base,
                "shortage_units": shortage_units,
                "excess_units": excess_units,
            }
        )

    transfer_rows: List[Dict[str, Any]] = []
    for description, items in transfer_pool.items():
        sources = [i for i in items if i.get("excess_units", 0) > 0.9]
        sinks = [i for i in items if i.get("shortage_units", 0) > 0.9]
        if not sources or not sinks:
            continue
        sources = sorted(sources, key=lambda x: (-safe_float(x.get("excess_units")), -safe_float(x.get("doi"))))
        sinks = sorted(sinks, key=lambda x: (-safe_float(x.get("shortage_units")), safe_float(x.get("doi"))))
        remaining_sources = [{**s} for s in sources]
        for sink in sinks:
            need = safe_float(sink.get("shortage_units"))
            if need <= 0:
                continue
            for source in remaining_sources:
                if source["store"] == sink["store"]:
                    continue
                available = safe_float(source.get("excess_units"))
                if available <= 0:
                    continue
                move_units = min(available, need)
                if move_units <= 0:
                    continue
                transfer_rows.append(
                    {
                        "description": description,
                        "brand": sink.get("brand"),
                        "category": sink.get("category"),
                        "supplier": sink.get("supplier"),
                        "from_store": source.get("store"),
                        "to_store": sink.get("store"),
                        "transfer_units": move_units,
                        "from_doi": source.get("doi"),
                        "to_doi": sink.get("doi"),
                        "target_qty_to": sink.get("target_qty"),
                        "stock_units_from": source.get("stock_units"),
                        "stock_units_to": sink.get("stock_units"),
                        "severity": "high" if move_units >= 3 else "medium",
                        "priority": "transfer",
                    }
                )
                source["excess_units"] = available - move_units
                need -= move_units
                if need <= 0:
                    break

    return {
        "store_sku_snapshot": sort_records(snapshot, ["store", "department", "category", "brand", "description"]),
        "doi_store_sku": sort_records(doi_rows, ["store", "description"]),
        "shortage_alerts": sort_records(shortage_rows, ["severity", "store", "description"]),
        "excess_alerts": sort_records(excess_rows, ["severity", "store", "description"]),
        "transfer_candidates": sort_records(transfer_rows, ["to_store", "from_store", "description"]),
    }


def process_merma(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando merma...")
    store_sheet = reader.find_sheet("Merma_StoreMonth")
    detail_sheet = reader.find_sheet("Merma_Detail")
    supplier_sheet = reader.find_sheet("Merma_SupplierMonth")

    outputs: Dict[str, Any] = {
        "monthly_business": [],
        "monthly_store": [],
        "detail_store_description": [],
        "detail_supplier": [],
    }

    if store_sheet:
        alias_map = {
            "store": ["store", "tienda"],
            "month_key": ["calmonth", "month", "mes"],
            "shrink_total": ["merma total", "merma"],
            "shrink_neg": ["merma neg", "negative shrink"],
            "sales": ["total sales", "sales", "ventas"],
            "shrink_pct_neg": ["% merma neg", "shrink pct neg"],
            "shrink_pct": ["% merma total", "shrink pct total"],
        }
        col_map = reader.resolve_columns(store_sheet, alias_map)
        month_business: Dict[str, Dict[str, float]] = defaultdict(metric_bucket_factory)
        monthly_store_rows: List[Dict[str, Any]] = []
        if all(k in col_map for k in ["store", "month_key", "shrink_total", "shrink_neg", "sales"]):
            for row in reader.iter_rows(store_sheet, col_map):
                month_key = nonempty_str(row.get("month_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                shrink_total = safe_float(row.get("shrink_total"))
                shrink_neg = safe_float(row.get("shrink_neg"))
                sales = safe_float(row.get("sales"))
                month_business[month_key]["shrink_amount"] += shrink_total
                month_business[month_key]["shrink_amount_neg"] += shrink_neg
                month_business[month_key]["sales"] += sales
                monthly_store_rows.append(
                    {
                        "month_key": month_key,
                        "store": store,
                        "store_name": store,
                        "shrink_amount": shrink_total,
                        "shrink_amount_neg": shrink_neg,
                        "sales": sales,
                        "shrink_pct": safe_div(shrink_total, sales),
                        "shrink_pct_neg": safe_div(shrink_neg, sales),
                    }
                )
            outputs["monthly_store"] = sort_records(monthly_store_rows, ["month_key", "store"])
            outputs["monthly_business"] = sort_records(
                [
                    {
                        "month_key": month_key,
                        "shrink_amount": vals.get("shrink_amount", 0.0),
                        "shrink_amount_neg": vals.get("shrink_amount_neg", 0.0),
                        "sales": vals.get("sales", 0.0),
                        "shrink_pct": safe_div(vals.get("shrink_amount", 0.0), vals.get("sales", 0.0)),
                        "shrink_pct_neg": safe_div(vals.get("shrink_amount_neg", 0.0), vals.get("sales", 0.0)),
                    }
                    for month_key, vals in month_business.items()
                ],
                ["month_key"],
            )
        else:
            ctx.warn(f"Hoja {store_sheet}: faltan columnas para merma mensual por tienda.")
    else:
        ctx.warn("No se encontró hoja Merma_StoreMonth.")

    if detail_sheet:
        alias_map = {
            "store": ["store", "tienda"],
            "month_key": ["calmonth", "month", "mes"],
            "supplier": ["supplier", "proveedor"],
            "category": ["category", "categoria"],
            "brand": ["brand", "marca"],
            "description": ["description", "descripcion"],
            "sales": ["total sales", "sales"],
            "shrink_amount": ["merma", "shrink"],
            "qty": ["quantity packs", "quantity (packs)", "qty"],
        }
        col_map = reader.resolve_columns(detail_sheet, alias_map)
        if all(k in col_map for k in ["store", "month_key", "description", "sales", "shrink_amount"]):
            detail_rows: List[Dict[str, Any]] = []
            for row in reader.iter_rows(detail_sheet, col_map):
                sales = safe_float(row.get("sales"))
                shrink_amount = safe_float(row.get("shrink_amount"))
                detail_rows.append(
                    {
                        "month_key": nonempty_str(row.get("month_key")),
                        "store": nonempty_str(row.get("store"), "UNKNOWN"),
                        "store_name": nonempty_str(row.get("store"), "UNKNOWN"),
                        "supplier": nonempty_str(row.get("supplier"), "Sin proveedor"),
                        "category": nonempty_str(row.get("category"), "Sin categoría"),
                        "brand": nonempty_str(row.get("brand"), "Sin marca"),
                        "description": nonempty_str(row.get("description"), "Sin descripción"),
                        "sales": sales,
                        "shrink_amount": shrink_amount,
                        "shrink_pct": safe_div(shrink_amount, sales),
                        "qty": safe_float(row.get("qty")),
                    }
                )
            outputs["detail_store_description"] = sort_records(detail_rows, ["month_key", "store", "description"])
        else:
            ctx.warn(f"Hoja {detail_sheet}: faltan columnas para merma detalle.")
    else:
        ctx.warn("No se encontró hoja Merma_Detail.")

    if supplier_sheet:
        alias_map = {
            "store": ["store", "tienda"],
            "month_key": ["calmonth", "month", "mes"],
            "supplier": ["supplier", "proveedor"],
            "sales": ["total sales", "sales"],
            "shrink_amount": ["merma", "shrink"],
            "shrink_neg": ["mermaneg", "merma neg", "negative shrink"],
            "shrink_pct": ["% merma total", "shrink pct total"],
            "shrink_pct_neg": ["% merma neg", "shrink pct neg"],
        }
        col_map = reader.resolve_columns(supplier_sheet, alias_map)
        if all(k in col_map for k in ["month_key", "supplier", "sales", "shrink_amount"]):
            agg: Dict[Tuple[str, str], Dict[str, Any]] = defaultdict(dict)
            for row in reader.iter_rows(supplier_sheet, col_map):
                month_key = nonempty_str(row.get("month_key"))
                supplier = nonempty_str(row.get("supplier"), "Sin proveedor")
                key = (month_key, supplier)
                bucket = agg.get(key) or {
                    "month_key": month_key,
                    "supplier": supplier,
                    "sales": 0.0,
                    "shrink_amount": 0.0,
                    "shrink_amount_neg": 0.0,
                    "stores": set(),
                }
                bucket["sales"] += safe_float(row.get("sales"))
                bucket["shrink_amount"] += safe_float(row.get("shrink_amount"))
                bucket["shrink_amount_neg"] += safe_float(row.get("shrink_neg"))
                bucket["stores"].add(nonempty_str(row.get("store"), "UNKNOWN"))
                agg[key] = bucket
            outputs["detail_supplier"] = sort_records(
                [
                    {
                        "month_key": v["month_key"],
                        "supplier": v["supplier"],
                        "store_count": len(v["stores"]),
                        "sales": v["sales"],
                        "shrink_amount": v["shrink_amount"],
                        "shrink_amount_neg": v["shrink_amount_neg"],
                        "shrink_pct": safe_div(v["shrink_amount"], v["sales"]),
                        "shrink_pct_neg": safe_div(v["shrink_amount_neg"], v["sales"]),
                    }
                    for v in agg.values()
                ],
                ["month_key", "supplier"],
            )
        else:
            ctx.warn(f"Hoja {supplier_sheet}: faltan columnas para merma por proveedor.")
    else:
        ctx.warn("No se encontró hoja Merma_SupplierMonth.")

    return outputs


def process_sbf(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando SBF...")
    weekly_sheet = reader.find_sheet("Cubo_SBF_Fact")
    monthly_sheet = reader.find_sheet("Chart_SBF_Month_XF")

    outputs = {
        "monthly_business": [],
        "weekly_business": [],
        "monthly_service_store": [],
        "weekly_service_store": [],
    }

    if weekly_sheet:
        alias_map = {
            "week_key": ["yearweek", "yearmonth", "week"],
            "store": ["store", "tienda"],
            "service": ["provider", "service", "servicio"],
            "hour": ["hour", "hora"],
            "tx": ["transactions", "tx", "transacciones"],
        }
        col_map = reader.resolve_columns(weekly_sheet, alias_map)
        if all(k in col_map for k in ["week_key", "store", "service", "tx"]):
            wbiz: Dict[str, float] = defaultdict(float)
            wsvc: Dict[Tuple[str, str, str], float] = defaultdict(float)
            for row in reader.iter_rows(weekly_sheet, col_map):
                week_key = nonempty_str(row.get("week_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                service = nonempty_str(row.get("service"), "Sin servicio")
                tx = safe_float(row.get("tx"))
                wbiz[week_key] += tx
                wsvc[(week_key, service, store)] += tx
            outputs["weekly_business"] = sort_records(
                apply_period_comparison(
                    [{"week_key": k, "tx": v} for k, v in wbiz.items()],
                    "week_key",
                    [],
                    ["tx"],
                    "week",
                ),
                ["week_key"],
            )
            outputs["weekly_service_store"] = sort_records(
                apply_period_comparison(
                    [{"week_key": k[0], "service": k[1], "store": k[2], "store_name": k[2], "tx": v} for k, v in wsvc.items()],
                    "week_key",
                    ["service", "store"],
                    ["tx"],
                    "week",
                ),
                ["week_key", "service", "store"],
            )
        else:
            ctx.warn(f"Hoja {weekly_sheet}: faltan columnas para SBF semanal.")
    else:
        ctx.warn("No se encontró hoja Cubo_SBF_Fact.")

    if monthly_sheet:
        alias_map = {
            "month_key": ["calmonth", "month", "mes"],
            "store": ["store", "tienda"],
            "service": ["provider", "service", "servicio"],
            "tx": ["tx", "transactions", "transacciones"],
        }
        col_map = reader.resolve_columns(monthly_sheet, alias_map)
        if all(k in col_map for k in ["month_key", "store", "service", "tx"]):
            mbiz: Dict[str, float] = defaultdict(float)
            msvc: Dict[Tuple[str, str, str], float] = defaultdict(float)
            for row in reader.iter_rows(monthly_sheet, col_map):
                month_key = nonempty_str(row.get("month_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                service = nonempty_str(row.get("service"), "Sin servicio")
                tx = safe_float(row.get("tx"))
                mbiz[month_key] += tx
                msvc[(month_key, service, store)] += tx
            outputs["monthly_business"] = sort_records(
                apply_period_comparison(
                    [{"month_key": k, "tx": v} for k, v in mbiz.items()],
                    "month_key",
                    [],
                    ["tx"],
                    "month",
                ),
                ["month_key"],
            )
            outputs["monthly_service_store"] = sort_records(
                apply_period_comparison(
                    [{"month_key": k[0], "service": k[1], "store": k[2], "store_name": k[2], "tx": v} for k, v in msvc.items()],
                    "month_key",
                    ["service", "store"],
                    ["tx"],
                    "month",
                ),
                ["month_key", "service", "store"],
            )
        else:
            ctx.warn(f"Hoja {monthly_sheet}: faltan columnas para SBF mensual.")
    else:
        ctx.warn("No se encontró hoja Chart_SBF_Month_XF.")

    return outputs


def process_cxc(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando CXC...")
    sheet = reader.find_sheet("CXC_DETAIL")
    if not sheet:
        ctx.warn("No se encontró hoja CXC_DETAIL. Se omite módulo CXC.")
        return {"summary": [], "detail": []}

    alias_map = {
        "store": ["store", "tienda"],
        "batch": ["batchnumber", "batch", "lote"],
        "time": ["time", "hora", "datetime"],
        "week_key": ["yearmonth", "yearweek", "week"],
        "week_num": ["yearweek", "weeknum", "semana"],
        "month_key": ["month", "calmonth", "mes"],
        "comment": ["comment", "comentario"],
        "cashier": ["cashier", "cajero"],
        "pos": ["pos"],
        "id": ["id", "reference", "referencia"],
        "amount": ["amount", "monto"],
    }
    col_map = reader.resolve_columns(sheet, alias_map)
    critical = ["store", "month_key", "amount"]
    missing = [c for c in critical if c not in col_map]
    if missing:
        ctx.warn(f"Hoja {sheet}: faltan columnas críticas para CXC: {missing}.")
        return {"summary": [], "detail": []}

    detail: List[Dict[str, Any]] = []
    summary_agg: Dict[Tuple[str, str], Dict[str, Any]] = defaultdict(dict)
    for row in reader.iter_rows(sheet, col_map):
        store = nonempty_str(row.get("store"), "UNKNOWN")
        month_key = nonempty_str(row.get("month_key"))
        week_key = nonempty_str(row.get("week_key"))
        amount = safe_float(row.get("amount"))
        rec = {
            "store": store,
            "store_name": store,
            "batch_number": nonempty_str(row.get("batch")),
            "time": row.get("time").isoformat() if isinstance(row.get("time"), (datetime, date)) else nonempty_str(row.get("time")),
            "month_key": month_key,
            "week_key": week_key,
            "comment": nonempty_str(row.get("comment")),
            "cashier": nonempty_str(row.get("cashier")),
            "pos": nonempty_str(row.get("pos")),
            "reference_id": nonempty_str(row.get("id")),
            "amount": amount,
        }
        detail.append(rec)
        skey = (store, month_key)
        bucket = summary_agg.get(skey) or {
            "store": store,
            "store_name": store,
            "month_key": month_key,
            "entry_count": 0,
            "amount": 0.0,
            "positive_amount": 0.0,
            "negative_amount": 0.0,
        }
        bucket["entry_count"] += 1
        bucket["amount"] += amount
        if amount >= 0:
            bucket["positive_amount"] += amount
        else:
            bucket["negative_amount"] += amount
        summary_agg[skey] = bucket

    summary = sort_records(list(summary_agg.values()), ["month_key", "store"])
    return {"summary": summary, "detail": sort_records(detail, ["month_key", "store", "time"])}


def process_prd0(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando PRD en cero...")
    detail_sheet = reader.find_sheet("PRD_CERO_DETAIL")
    if not detail_sheet:
        ctx.warn("No se encontró hoja PRD_CERO_DETAIL. Se omite módulo PRD en cero.")
        return {"supplier_summary": [], "supplier_detail": []}

    alias_map = {
        "store": ["store", "tienda"],
        "supplier": ["supplier", "proveedor"],
        "description": ["description", "descripcion"],
        "department": ["department", "departamento"],
        "category": ["category", "categoria"],
        "brand": ["brand", "marca"],
        "lookup_code": ["lookup code", "codigo", "sku"],
        "quantity": ["quantity", "qty"],
        "rop": ["rop"],
        "inactive_purchase": ["inactive compra", "inactive purchase"],
        "qty_sold_s4": ["qty sold s4", "qty s4"],
        "sales_s4": ["sales s4", "ventas s4"],
        "avg_daily_qty_s4": ["avg daily qty s4", "average daily qty s4"],
        "avg_daily_sales_s4": ["avg daily sales s4", "average daily sales s4"],
        "lost_units_3d": ["lost units 3d", "unidades perdidas 3d"],
        "lost_sales_3d": ["lost sales 3d", "ventas perdidas 3d"],
    }
    col_map = reader.resolve_columns(detail_sheet, alias_map)
    critical = ["store", "supplier", "description"]
    missing = [c for c in critical if c not in col_map]
    if missing:
        ctx.warn(f"Hoja {detail_sheet}: faltan columnas críticas para PRD en cero: {missing}")
        return {"supplier_summary": [], "supplier_detail": []}

    details: List[Dict[str, Any]] = []
    summary_agg: Dict[str, Dict[str, Any]] = defaultdict(dict)
    for row in reader.iter_rows(detail_sheet, col_map):
        supplier = nonempty_str(row.get("supplier"), "Sin proveedor")
        store = nonempty_str(row.get("store"), "UNKNOWN")
        quantity = safe_float(row.get("quantity"))
        rop = safe_float(row.get("rop"))
        lost_sales_3d = safe_float(row.get("lost_sales_3d"))
        rec = {
            "store": store,
            "store_name": store,
            "supplier": supplier,
            "department": nonempty_str(row.get("department"), "Sin departamento"),
            "category": nonempty_str(row.get("category"), "Sin categoría"),
            "brand": nonempty_str(row.get("brand"), "Sin marca"),
            "description": nonempty_str(row.get("description"), "Sin descripción"),
            "lookup_code": nonempty_str(row.get("lookup_code")),
            "stock_units": quantity,
            "rop": rop,
            "inactive_purchase": nonempty_str(row.get("inactive_purchase")),
            "qty_sold_s4": safe_float(row.get("qty_sold_s4")),
            "sales_s4": safe_float(row.get("sales_s4")),
            "avg_daily_qty_s4": safe_float(row.get("avg_daily_qty_s4")),
            "avg_daily_sales_s4": safe_float(row.get("avg_daily_sales_s4")),
            "lost_units_3d": safe_float(row.get("lost_units_3d")),
            "lost_sales_3d": lost_sales_3d,
            "priority": "high" if lost_sales_3d >= 5 else "medium" if lost_sales_3d > 0 else "low",
        }
        details.append(rec)
        bucket = summary_agg.get(supplier) or {
            "supplier": supplier,
            "item_count": 0,
            "store_count": set(),
            "qty_zero_count": 0,
            "lost_sales_3d": 0.0,
            "sales_s4": 0.0,
        }
        bucket["item_count"] += 1
        if quantity <= 0:
            bucket["qty_zero_count"] += 1
        bucket["store_count"].add(store)
        bucket["lost_sales_3d"] += lost_sales_3d
        bucket["sales_s4"] += safe_float(row.get("sales_s4"))
        summary_agg[supplier] = bucket

    summary = []
    for supplier, v in summary_agg.items():
        summary.append(
            {
                "supplier": supplier,
                "item_count": v["item_count"],
                "store_count": len(v["store_count"]),
                "qty_zero_count": v["qty_zero_count"],
                "lost_sales_3d": v["lost_sales_3d"],
                "sales_s4": v["sales_s4"],
                "priority": "high" if v["lost_sales_3d"] >= 20 else "medium" if v["lost_sales_3d"] >= 5 else "low",
            }
        )

    return {
        "supplier_summary": sort_records(summary, ["priority", "supplier"]),
        "supplier_detail": sort_records(details, ["supplier", "store", "description"]),
    }


def process_innovation(ctx: ExportContext, reader: WorkbookReader) -> Dict[str, Any]:
    Logger.info("Procesando innovation...")
    agg_sheet = reader.find_sheet("Innovation_Combos_Agg")
    detail_sheet = reader.find_sheet("Innovation_Combos_Detail")
    outputs = {
        "monthly_summary": [],
        "weekly_summary": [],
        "by_hour": [],
        "by_dow": [],
        "by_store": [],
    }

    if agg_sheet:
        alias_map = {
            "period_type": ["periodtype", "period type"],
            "period_key": ["periodkey", "period key", "month", "week"],
            "store": ["store", "tienda"],
            "combo_key": ["combokey", "combo key"],
            "combo_label": ["combolabel", "combo label"],
            "combo_short": ["comboshortlabel", "combo short label"],
            "qty": ["qty sold", "qty"],
            "sales": ["sales", "ventas"],
        }
        col_map = reader.resolve_columns(agg_sheet, alias_map)
        if all(k in col_map for k in ["period_type", "period_key", "store", "combo_label", "qty", "sales"]):
            monthly, weekly, store_agg = [], [], defaultdict(lambda: {"sales": 0.0, "qty": 0.0})
            for row in reader.iter_rows(agg_sheet, col_map):
                period_type = normalize_text(row.get("period_type"))
                period_key = nonempty_str(row.get("period_key"))
                store = nonempty_str(row.get("store"), "UNKNOWN")
                combo_key = nonempty_str(row.get("combo_key"), nonempty_str(row.get("combo_label"), "combo"))
                combo_label = nonempty_str(row.get("combo_label"), "Sin combo")
                combo_short = nonempty_str(row.get("combo_short"), combo_label)
                qty = safe_float(row.get("qty"))
                sales = safe_float(row.get("sales"))
                rec = {
                    "period_key": period_key,
                    "store": store,
                    "store_name": store,
                    "combo_key": combo_key,
                    "combo_label": combo_label,
                    "combo_short_label": combo_short,
                    "qty": qty,
                    "sales": sales,
                }
                store_agg[(period_type, period_key, store)]["sales"] += sales
                store_agg[(period_type, period_key, store)]["qty"] += qty
                if period_type == "month":
                    monthly.append({"month_key": period_key, **rec})
                elif period_type == "week":
                    weekly.append({"week_key": period_key, **rec})
            outputs["monthly_summary"] = sort_records(monthly, ["month_key", "store", "combo_label"])
            outputs["weekly_summary"] = sort_records(weekly, ["week_key", "store", "combo_label"])
            outputs["by_store"] = sort_records(
                [
                    {
                        "period_type": ptype,
                        "period_key": pkey,
                        "store": store,
                        "store_name": store,
                        "sales": vals["sales"],
                        "qty": vals["qty"],
                    }
                    for (ptype, pkey, store), vals in store_agg.items()
                ],
                ["period_type", "period_key", "store"],
            )
        else:
            ctx.warn(f"Hoja {agg_sheet}: faltan columnas para innovation aggregate.")
    else:
        ctx.warn("No se encontró hoja Innovation_Combos_Agg.")

    if detail_sheet:
        alias_map = {
            "week_key": ["yearweek", "week"],
            "month_key": ["calmonth", "month", "mes"],
            "store": ["store", "tienda"],
            "combo_key": ["combokey", "combo key"],
            "combo_label": ["combolabel", "combo label"],
            "combo_short": ["comboshortlabel", "combo short label"],
            "dow": ["dow", "dayofweek", "dia"],
            "dow_name": ["dow name", "dow_name", "weekday"],
            "qty": ["qty sold", "qty"],
            "sales": ["sales", "ventas"],
        }
        col_map = reader.resolve_columns(detail_sheet, alias_map)
        if all(k in col_map for k in ["store", "combo_label", "sales"]):
            by_dow: Dict[Tuple[str, str, str, str, int], Dict[str, Any]] = defaultdict(dict)
            for row in reader.iter_rows(detail_sheet, col_map):
                store = nonempty_str(row.get("store"), "UNKNOWN")
                combo_key = nonempty_str(row.get("combo_key"), nonempty_str(row.get("combo_label"), "combo"))
                combo_label = nonempty_str(row.get("combo_label"), "Sin combo")
                combo_short = nonempty_str(row.get("combo_short"), combo_label)
                dow = safe_int(row.get("dow"), 0)
                dow_name = nonempty_str(row.get("dow_name"), f"DOW {dow}")
                dow_group = make_dow_group(dow_name, dow)
                qty = safe_float(row.get("qty"))
                sales = safe_float(row.get("sales"))
                month_key = nonempty_str(row.get("month_key"))
                week_key = nonempty_str(row.get("week_key"))
                if month_key:
                    key = ("month", month_key, store, combo_key, dow)
                    bucket = by_dow.get(key) or {
                        "period_type": "month",
                        "period_key": month_key,
                        "store": store,
                        "store_name": store,
                        "combo_key": combo_key,
                        "combo_label": combo_label,
                        "combo_short_label": combo_short,
                        "dow": dow,
                        "dow_name": dow_name,
                        "dow_group": dow_group,
                        "qty": 0.0,
                        "sales": 0.0,
                    }
                    bucket["qty"] += qty
                    bucket["sales"] += sales
                    by_dow[key] = bucket
                if week_key:
                    key = ("week", week_key, store, combo_key, dow)
                    bucket = by_dow.get(key) or {
                        "period_type": "week",
                        "period_key": week_key,
                        "store": store,
                        "store_name": store,
                        "combo_key": combo_key,
                        "combo_label": combo_label,
                        "combo_short_label": combo_short,
                        "dow": dow,
                        "dow_name": dow_name,
                        "dow_group": dow_group,
                        "qty": 0.0,
                        "sales": 0.0,
                    }
                    bucket["qty"] += qty
                    bucket["sales"] += sales
                    by_dow[key] = bucket
            outputs["by_dow"] = sort_records(list(by_dow.values()), ["period_type", "period_key", "store", "combo_key", "dow"])
        else:
            ctx.warn(f"Hoja {detail_sheet}: faltan columnas para innovation by_dow.")
    else:
        ctx.warn("No se encontró hoja Innovation_Combos_Detail.")

    ctx.warn("Innovation by_hour no pudo generarse: no se detectó columna horaria en la fuente de innovation.")
    outputs["by_hour"] = []
    return outputs


# -----------------------------
# Hallazgos signals
# -----------------------------
def build_sales_signals(sales_outputs: Dict[str, Any]) -> List[Dict[str, Any]]:
    signals: List[Dict[str, Any]] = []
    monthly_department = sales_outputs.get("monthly_department", [])
    monthly_store = sales_outputs.get("monthly_store", [])
    if monthly_department:
        latest_month = max((r.get("month_key") for r in monthly_department if r.get("month_key")), default=None)
        latest_rows = [r for r in monthly_department if r.get("month_key") == latest_month]
        latest_rows = sorted(latest_rows, key=lambda x: safe_float(x.get("sales_delta_pct")))
        for row in latest_rows[:12] + latest_rows[-12:]:
            delta_pct = safe_float(row.get("sales_delta_pct"))
            current = safe_float(row.get("sales"))
            compare = safe_float(row.get("sales_ly"))
            signals.append(
                {
                    "domain": "sales",
                    "level": "department",
                    "scope_key": row.get("department"),
                    "period_key": latest_month,
                    "metric": "sales",
                    "current_value": current,
                    "compare_value": compare,
                    "delta": current - compare,
                    "delta_pct": delta_pct,
                    "severity": severity_from_pct(delta_pct, reverse=True),
                    "suggested_focus": "Acelerar activaciones/combo y revisar mix" if delta_pct < 0 else "Escalar dinámica ganadora y asegurar abastecimiento",
                }
            )
    if monthly_store:
        latest_month = max((r.get("month_key") for r in monthly_store if r.get("month_key")), default=None)
        for row in [r for r in monthly_store if r.get("month_key") == latest_month]:
            delta_pct = safe_float(row.get("sales_delta_pct"))
            current = safe_float(row.get("sales"))
            compare = safe_float(row.get("sales_ly"))
            signals.append(
                {
                    "domain": "sales",
                    "level": "store",
                    "scope_key": row.get("store"),
                    "period_key": latest_month,
                    "metric": "sales",
                    "current_value": current,
                    "compare_value": compare,
                    "delta": current - compare,
                    "delta_pct": delta_pct,
                    "severity": severity_from_pct(delta_pct, reverse=True),
                    "suggested_focus": "Profundizar revisión tienda y fricción operativa" if delta_pct < 0 else "Replicar ejecución comercial de tienda referente",
                }
            )
    return sort_records(signals, ["period_key", "level", "scope_key"])


def build_delivery_signals(delivery_outputs: Dict[str, Any]) -> List[Dict[str, Any]]:
    signals: List[Dict[str, Any]] = []
    rows = delivery_outputs.get("monthly_store_platform", [])
    latest_month = max((r.get("month_key") for r in rows if r.get("month_key")), default=None)
    latest_rows = [r for r in rows if r.get("month_key") == latest_month]
    latest_rows = sorted(latest_rows, key=lambda x: safe_float(x.get("sales_delta_pct")))
    for row in latest_rows[:10] + latest_rows[-10:]:
        delta_pct = safe_float(row.get("sales_delta_pct"))
        current = safe_float(row.get("sales"))
        compare = safe_float(row.get("sales_ly"))
        signals.append(
            {
                "domain": "delivery",
                "level": "store_platform",
                "scope_key": f"{row.get('store')}|{row.get('platform')}",
                "period_key": latest_month,
                "metric": "sales",
                "current_value": current,
                "compare_value": compare,
                "delta": current - compare,
                "delta_pct": delta_pct,
                "severity": severity_from_pct(delta_pct, reverse=True),
                "suggested_focus": "Revisar surtido, disponibilidad y ranking en app" if delta_pct < 0 else "Aumentar pauta y proteger fill-rate del canal",
            }
        )
    return sort_records(signals, ["period_key", "scope_key"])


def build_inventory_signals(inventory_outputs: Dict[str, Any]) -> List[Dict[str, Any]]:
    signals: List[Dict[str, Any]] = []
    for row in inventory_outputs.get("shortage_alerts", [])[:80]:
        signals.append(
            {
                "domain": "inventory",
                "level": "store_sku",
                "scope_key": f"{row.get('store')}|{row.get('description')}",
                "period_key": "snapshot",
                "metric": "shortage_units",
                "current_value": safe_float(row.get("stock_units")),
                "compare_value": safe_float(row.get("target_qty")),
                "delta": safe_float(row.get("stock_units")) - safe_float(row.get("target_qty")),
                "delta_pct": safe_div(safe_float(row.get("stock_units")) - safe_float(row.get("target_qty")), safe_float(row.get("target_qty"))),
                "severity": row.get("severity"),
                "suggested_focus": "Reabastecer o transferir inventario de otra tienda",
            }
        )
    for row in inventory_outputs.get("excess_alerts", [])[:80]:
        signals.append(
            {
                "domain": "inventory",
                "level": "store_sku",
                "scope_key": f"{row.get('store')}|{row.get('description')}",
                "period_key": "snapshot",
                "metric": "doi",
                "current_value": safe_float(row.get("doi")),
                "compare_value": safe_float(row.get("doi_ideal")),
                "delta": safe_float(row.get("doi")) - safe_float(row.get("doi_ideal")),
                "delta_pct": safe_div(safe_float(row.get("doi")) - safe_float(row.get("doi_ideal")), safe_float(row.get("doi_ideal"))),
                "severity": row.get("severity"),
                "suggested_focus": "Reducir compra, promover salida o evaluar traslado",
            }
        )
    return sort_records(signals, ["severity", "scope_key"])


def build_shrink_signals(merma_outputs: Dict[str, Any]) -> List[Dict[str, Any]]:
    signals: List[Dict[str, Any]] = []
    rows = merma_outputs.get("detail_store_description", [])
    if not rows:
        return signals
    latest_month = max((r.get("month_key") for r in rows if r.get("month_key")), default=None)
    latest_rows = [r for r in rows if r.get("month_key") == latest_month]
    latest_rows = sorted(latest_rows, key=lambda x: safe_float(x.get("shrink_amount")))
    for row in latest_rows[:80]:
        shrink_amount = safe_float(row.get("shrink_amount"))
        signals.append(
            {
                "domain": "shrink",
                "level": "store_description",
                "scope_key": f"{row.get('store')}|{row.get('description')}",
                "period_key": latest_month,
                "metric": "shrink_amount",
                "current_value": shrink_amount,
                "compare_value": safe_float(row.get("sales")),
                "delta": shrink_amount,
                "delta_pct": safe_float(row.get("shrink_pct")),
                "severity": "high" if shrink_amount <= -20 or safe_float(row.get("shrink_pct")) <= -0.03 else "medium",
                "suggested_focus": "Auditar causa raíz, manipulación y merma operativa",
            }
        )
    return sort_records(signals, ["period_key", "scope_key"])


# -----------------------------
# Meta builders
# -----------------------------
def build_calendars(months: Sequence[str], weeks: Sequence[str]) -> Dict[str, Any]:
    return {
        "months": [
            {
                "month_key": month,
                "comparable_ly_month": ly_month_key(month),
                "lm_month": prev_month_key(month),
            }
            for month in sorted(set(m for m in months if m))
        ],
        "weeks": [
            {
                "week_key": week,
                "comparable_ly_week": ly_week_key(week),
            }
            for week in sorted(set(w for w in weeks if w))
        ],
    }


def build_stores_json(stores: Sequence[str]) -> List[Dict[str, Any]]:
    return [{"store": store, "store_name": store} for store in sorted(set(s for s in stores if s))]


def merge_catalogs(*catalog_sets: Dict[str, Any]) -> Dict[str, Any]:
    out = {
        "departments": set(),
        "categories": set(),
        "brands": set(),
        "descriptions": set(),
        "delivery_platforms": set(),
        "sbf_services": set(),
        "suppliers": set(),
    }
    for cat in catalog_sets:
        for key in out:
            vals = cat.get(key, []) if isinstance(cat, dict) else []
            out[key].update(v for v in vals if v)
    return {k: sorted(v) for k, v in out.items()}


def collect_sbf_services(reader: WorkbookReader) -> List[str]:
    sheet = reader.find_sheet("Cubo_SBF_Fact") or reader.find_sheet("Chart_SBF_Month_XF")
    if not sheet:
        return []
    alias_map = {"service": ["provider", "service", "servicio"]}
    col_map = reader.resolve_columns(sheet, alias_map)
    if "service" not in col_map:
        return []
    services = set()
    for row in reader.iter_rows(sheet, col_map):
        service = nonempty_str(row.get("service"))
        if service:
            services.add(service)
    return sorted(services)


# -----------------------------
# Export orchestration
# -----------------------------
def export_empty_with_warning(ctx: ExportContext, relative_path: str, warning: str) -> None:
    ctx.warn(warning)
    write_json(ctx, relative_path, [])


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    source_file = project_root / SOURCE_FILENAME
    out_root = project_root / "data_api" / "out"
    ensure_dirs(out_root)
    ctx = ExportContext(project_root=project_root, source_file=source_file, out_root=out_root)

    Logger.info(f"Iniciando exportador maestro JSON v{VERSION_EXPORTER}")
    Logger.info(f"Proyecto: {project_root}")
    Logger.info(f"Fuente: {source_file}")

    if not source_file.exists():
        Logger.error(f"No se encontró el archivo fuente: {source_file}")
        return 1

    reader = WorkbookReader(source_file)
    ctx.detected_sheets = list(reader.sheetnames)
    Logger.info(f"Hojas detectadas: {len(ctx.detected_sheets)}")

    sales_results = process_sales_and_delivery(ctx, reader)
    merge_tx_into_sales_outputs(ctx, reader, sales_results)
    inventory_results = process_inventory(ctx, reader, sales_results)
    merma_results = process_merma(ctx, reader)
    sbf_results = process_sbf(ctx, reader)
    cxc_results = process_cxc(ctx, reader)
    prd0_results = process_prd0(ctx, reader)
    innovation_results = process_innovation(ctx, reader)

    catalogs = merge_catalogs(
        sales_results.get("catalogs", {}),
        {
            "suppliers": [r.get("supplier") for r in inventory_results.get("store_sku_snapshot", [])],
            "descriptions": [r.get("description") for r in inventory_results.get("store_sku_snapshot", [])],
            "brands": [r.get("brand") for r in inventory_results.get("store_sku_snapshot", [])],
            "categories": [r.get("category") for r in inventory_results.get("store_sku_snapshot", [])],
            "departments": [r.get("department") for r in inventory_results.get("store_sku_snapshot", [])],
            "delivery_platforms": sales_results.get("catalogs", {}).get("delivery_platforms", []),
            "sbf_services": collect_sbf_services(reader),
        },
    )
    stores = sales_results.get("catalogs", {}).get("stores", []) or sorted({r.get("store") for r in inventory_results.get("store_sku_snapshot", []) if r.get("store")})
    months = sales_results.get("catalogs", {}).get("months", []) or sorted({r.get("month_key") for r in merma_results.get("monthly_business", []) if r.get("month_key")})
    weeks = sales_results.get("catalogs", {}).get("weeks", []) or sorted({r.get("week_key") for r in sbf_results.get("weekly_business", []) if r.get("week_key")})

    ctx.stores_available = stores
    ctx.months_available = months
    ctx.weeks_available = weeks

    # Meta
    write_json(ctx, "meta/stores.json", build_stores_json(stores))
    write_json(ctx, "meta/calendars.json", build_calendars(months, weeks))
    write_json(ctx, "meta/catalogs.json", catalogs)

    # Ventas
    sales_outputs = sales_results.get("sales_outputs", {})
    for name in [
        "monthly_business",
        "weekly_business",
        "monthly_store",
        "weekly_store",
        "monthly_department",
        "weekly_department",
        "monthly_category",
        "weekly_category",
        "monthly_brand",
        "weekly_brand",
        "monthly_description",
        "weekly_description",
        "sales_by_dow",
    ]:
        write_json(ctx, f"ventas/{name}.json", json_ready_records(sales_outputs.get(name, [])))
    export_empty_with_warning(ctx, "ventas/sales_by_hour.json", "Ventas por hora no pudo generarse: la fuente no incluye columna hora para ventas.")

    # Delivery
    delivery_outputs = sales_results.get("delivery_outputs", {})
    for name in [
        "monthly_business",
        "weekly_business",
        "monthly_store_platform",
        "weekly_store_platform",
        "monthly_category",
        "weekly_category",
        "delivery_by_dow",
    ]:
        write_json(ctx, f"delivery/{name}.json", json_ready_records(delivery_outputs.get(name, [])))
    export_empty_with_warning(ctx, "delivery/delivery_by_hour.json", "Delivery por hora no pudo generarse: la fuente no incluye columna hora para delivery sales.")

    # Inventario
    for name in ["store_sku_snapshot", "doi_store_sku", "shortage_alerts", "excess_alerts", "transfer_candidates"]:
        write_json(ctx, f"inventario/{name}.json", json_ready_records(inventory_results.get(name, [])))

    # Merma
    for name in ["monthly_business", "monthly_store", "detail_store_description", "detail_supplier"]:
        write_json(ctx, f"merma/{name}.json", json_ready_records(merma_results.get(name, [])))

    # SBF
    for name in ["monthly_business", "weekly_business", "monthly_service_store", "weekly_service_store"]:
        write_json(ctx, f"sbf/{name}.json", json_ready_records(sbf_results.get(name, [])))

    # CXC
    for name in ["summary", "detail"]:
        write_json(ctx, f"cxc/{name}.json", json_ready_records(cxc_results.get(name, [])))

    # PRD0
    for name in ["supplier_summary", "supplier_detail"]:
        write_json(ctx, f"prd0/{name}.json", json_ready_records(prd0_results.get(name, [])))

    # Hallazgos signals
    hallazgos = {
        "sales_signals": build_sales_signals(sales_outputs),
        "delivery_signals": build_delivery_signals(delivery_outputs),
        "inventory_signals": build_inventory_signals(inventory_results),
        "shrink_signals": build_shrink_signals(merma_results),
    }
    for name, payload in hallazgos.items():
        write_json(ctx, f"hallazgos/{name}.json", json_ready_records(payload))

    # Innovation
    for name in ["monthly_summary", "weekly_summary", "by_hour", "by_dow", "by_store"]:
        write_json(ctx, f"innovation/{name}.json", json_ready_records(innovation_results.get(name, [])))

    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_file": ctx.source_file.name,
        "detected_sheets": ctx.detected_sheets,
        "months_available": ctx.months_available,
        "weeks_available": ctx.weeks_available,
        "files_generated": ctx.files_generated,
        "row_counts_by_file": ctx.row_counts_by_file,
        "warnings": ctx.warnings,
        "version_exporter": VERSION_EXPORTER,
    }
    write_json(ctx, "meta/manifest.json", manifest)

    Logger.info("Exportación completada con éxito.")
    Logger.info(f"Archivos generados: {len(ctx.files_generated)}")
    if ctx.warnings:
        Logger.warn(f"Warnings detectados: {len(ctx.warnings)}. Revisar data_api/out/meta/manifest.json")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as exc:
        Logger.error(f"Fallo no controlado: {exc}")
        traceback.print_exc()
        sys.exit(1)
