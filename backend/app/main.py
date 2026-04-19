from __future__ import annotations

import io
import math
import os
import re
import threading
import time
import unicodedata
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

APP_VERSION = "chat-data-engine-v2"
CACHE_TTL_SECONDS = int(os.getenv("WORKBOOK_CACHE_TTL_SECONDS", "900"))
DEFAULT_DATA_URL = os.getenv("WORKBOOK_PUBLIC_URL", "").strip()
REQUEST_TIMEOUT = int(os.getenv("WORKBOOK_REQUEST_TIMEOUT_SECONDS", "90"))

MONTHS_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
MONTHS_ABBR = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun", 7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}
DOW_ES = {"lunes": 1, "martes": 2, "miercoles": 3, "miércoles": 3, "jueves": 4, "viernes": 5, "sabado": 6, "sábado": 6, "domingo": 7}

SHEETS: Dict[str, Tuple[str, List[str]]] = {
    "sales": (
        "Cubo_Sales_DetailHallazgos",
        ["Date Sold", "YearWeek", "CalMonth", "Store", "Department", "Category", "Brand", "Description", "Daypart", "DOW", "DOW_Name", "Is_Delivery", "Delivery_Channel", "Qty Sold", "Sales", "Total Gross Margin"],
    ),
    "sbf_month": ("Chart_SBF_Month_XF", ["CalMonth", "Store", "Provider", "Daypart", "DOW", "Hour", "TX"]),
    "inventory": ("Inventario_Detail", ["Store", "Department", "Supplier", "Category", "Brand", "Description", "Amount", "Quantity", "Días Inv", "Qty Sold S4", "Sales S4", "IdealQty", "TargetQty"]),
    "merma": ("Merma_Detail", ["Store", "CalMonth", "Supplier", "Category", "Brand", "Description", "Total Sales", "Merma", "Quantity (packs)"]),
    "innovation": (
        "Innovation_Combos_Detail",
        ["Date Sold", "YearWeek", "CalMonth", "Store", "Department", "Category", "Brand", "Description", "ComboLabel", "Daypart", "DOW", "DOW_Name", "Is_Delivery", "Delivery_Channel", "Qty Sold", "Sales", "Total Gross Margin"],
    ),
    "cxc": ("CXC_DETAIL", ["Store", "Time", "Month", "Comment", "Cashier", "Amount"]),
    "prd0": ("PRD_CERO_DETAIL", ["Store", "Supplier", "Description", "Department", "Category", "Brand", "Quantity", "ROP", "Qty Sold S4", "Sales S4", "Lost Units 3d", "Lost Sales 3d"]),
}
TEXT_COLUMNS = {
    "sales": ["Store", "Department", "Category", "Brand", "Description", "Daypart", "DOW_Name", "Delivery_Channel"],
    "sbf_month": ["Store", "Provider", "Daypart"],
    "inventory": ["Store", "Department", "Supplier", "Category", "Brand", "Description"],
    "merma": ["Store", "Supplier", "Category", "Brand", "Description"],
    "innovation": ["Store", "Department", "Category", "Brand", "Description", "ComboLabel", "Daypart", "DOW_Name", "Delivery_Channel"],
    "cxc": ["Store", "Comment", "Cashier"],
    "prd0": ["Store", "Supplier", "Description", "Department", "Category", "Brand"],
}
DOMAIN_LABELS = {"sales": "Ventas", "delivery": "Delivery", "sbf": "SBF", "inventory": "Inventario", "merma": "Merma", "innovation": "Innovación", "cxc": "CXC", "prd0": "PRD en cero", "hallazgos": "Hallazgos"}
GROUP_ALIASES = {
    "tienda": "Store", "store": "Store", "categoria": "Category", "category": "Category", "marca": "Brand", "brand": "Brand",
    "descripcion": "Description", "description": "Description", "departamento": "Department", "department": "Department",
    "proveedor": "Provider", "provider": "Provider", "servicio": "Provider", "canal": "Delivery_Channel", "channel": "Delivery_Channel",
    "dia de semana": "DOW_Name", "dow": "DOW_Name", "dia": "Date Sold", "fecha": "Date Sold",
    "horario": "Daypart", "daypart": "Daypart", "turno": "Daypart", "hora": "Hour", "combo": "ComboLabel",
}

class QueryRequest(BaseModel):
    question: str
    filters: Dict[str, Any] = Field(default_factory=dict)
    page_context: Dict[str, Any] = Field(default_factory=dict)
    data_url: Optional[str] = None

@dataclass
class SourceCache:
    source: str
    loaded_at: float
    workbook_bytes: bytes
    frames: Dict[str, pd.DataFrame] = field(default_factory=dict)
    lock: threading.Lock = field(default_factory=threading.Lock)

_cache: Dict[str, SourceCache] = {}
_cache_lock = threading.Lock()

app = FastAPI(title="AMPM Analytical Chat Engine", version=APP_VERSION)
origins_raw = os.getenv("ALLOWED_ORIGINS", "*").strip()
allow_origins = [o.strip() for o in origins_raw.split(",") if o.strip()] or ["*"]
app.add_middleware(CORSMiddleware, allow_origins=allow_origins, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


def norm(value: Any) -> str:
    raw = "" if value is None else str(value)
    raw = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip()


def txt(value: Any) -> str:
    return "" if value is None else str(value).strip()


def fmt_money(value: float) -> str:
    return f"${float(value):,.2f}"


def fmt_num(value: float) -> str:
    return f"{float(value):,.0f}"


def fmt_pct(value: Optional[float]) -> str:
    if value is None:
        return "N/D"
    try:
        val = float(value)
    except Exception:
        return "N/D"
    if not math.isfinite(val):
        return "N/D"
    return f"{val * 100:.1f}%"


def pct_delta(cur: float, prev: float) -> Optional[float]:
    if prev in (None, 0) or pd.isna(prev):
        return None
    return (float(cur) / float(prev)) - 1.0


def month_label(month_key: str) -> str:
    try:
        year, month = month_key.split("-")
        return f"{MONTHS_ABBR[int(month)]} {year}"
    except Exception:
        return month_key


def same_month_ly(month_key: str) -> Optional[str]:
    try:
        year, month = month_key.split("-")
        return f"{int(year)-1:04d}-{month}"
    except Exception:
        return None


def read_sheet_df(ws: openpyxl.worksheet.worksheet.Worksheet, columns: List[str]) -> pd.DataFrame:
    rows = ws.iter_rows(values_only=True)
    header = ["" if c is None else str(c).strip() for c in next(rows)]
    positions = [header.index(col) for col in columns]
    data: List[List[Any]] = []
    for row in rows:
        data.append([row[pos] if pos < len(row) else None for pos in positions])
    return pd.DataFrame(data, columns=columns)


def prepare_frame(name: str, df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if name in {"sales", "innovation"} and "Date Sold" in df.columns:
        df["Date Sold"] = pd.to_datetime(df["Date Sold"], errors="coerce")
    if name == "cxc" and "Time" in df.columns:
        df["Time"] = pd.to_datetime(df["Time"], errors="coerce")
    for col in ["CalMonth", "YearWeek", "Month"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)
    if "Is_Delivery" in df.columns:
        df["Is_Delivery"] = df["Is_Delivery"].apply(lambda v: True if v in (True, 1, "1", "true", "True", "TRUE") else False)
    for col in TEXT_COLUMNS.get(name, []):
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str)
            df[f"_{col}"] = df[col].map(norm)
    for col in ["Qty Sold", "Sales", "Total Gross Margin", "TX", "Amount", "Quantity", "Días Inv", "Qty Sold S4", "Sales S4", "IdealQty", "TargetQty", "Merma", "Total Sales", "Quantity (packs)", "ROP", "Lost Units 3d", "Lost Sales 3d", "Hour", "DOW"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def get_source_cache(source: str) -> SourceCache:
    now = time.time()
    with _cache_lock:
        cached = _cache.get(source)
        if cached and now - cached.loaded_at <= CACHE_TTL_SECONDS:
            return cached
    if source.startswith("http://") or source.startswith("https://"):
        resp = requests.get(source, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        workbook_bytes = resp.content
    else:
        if not os.path.exists(source):
            raise FileNotFoundError(source)
        with open(source, "rb") as fh:
            workbook_bytes = fh.read()
    cache = SourceCache(source=source, loaded_at=now, workbook_bytes=workbook_bytes)
    with _cache_lock:
        _cache[source] = cache
    return cache


def get_frame(cache: SourceCache, key: str) -> pd.DataFrame:
    if key in cache.frames:
        return cache.frames[key]
    with cache.lock:
        if key in cache.frames:
            return cache.frames[key]
        sheet_name, columns = SHEETS[key]
        wb = openpyxl.load_workbook(io.BytesIO(cache.workbook_bytes), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"No encontré la hoja requerida: {sheet_name}")
        df = prepare_frame(key, read_sheet_df(wb[sheet_name], columns))
        cache.frames[key] = df
        return df


def available_months(df: pd.DataFrame) -> List[str]:
    col = "CalMonth" if "CalMonth" in df.columns else "Month" if "Month" in df.columns else None
    if not col:
        return []
    return sorted([m for m in df[col].dropna().astype(str).unique().tolist() if re.match(r"^\d{4}-\d{2}$", str(m))])


def extract_stores(qn: str, filters: Dict[str, Any], sales_df: pd.DataFrame) -> List[str]:
    stores = []
    for match in re.findall(r"\b(?:ampm\s*0?([1-9])|a0?([1-9]))\b", qn):
        digit = next((m for m in match if m), None)
        if digit:
            stores.append(f"AMPM{int(digit):02d}")
    if not stores:
        for raw in filters.get("stores") or []:
            val = txt(raw).upper()
            if val:
                stores.append(val)
    valid = set(sales_df["Store"].dropna().astype(str).unique().tolist())
    return sorted(set([s for s in stores if s in valid]))


def extract_months(qn: str, filters: Dict[str, Any], months: List[str]) -> List[str]:
    explicit: List[str] = []
    for name, month_num in MONTHS_ES.items():
        for match in re.finditer(rf"\b{name}\b(?:\s+de)?\s*(20\d{{2}})?", qn):
            year = match.group(1)
            if year:
                explicit.append(f"{year}-{month_num:02d}")
            else:
                candidates = [m for m in months if m.endswith(f"-{month_num:02d}")]
                if candidates:
                    explicit.append(candidates[-1])
    if explicit:
        return [m for m in sorted(set(explicit)) if m in months]
    for y, m in re.findall(r"\b(20\d{2})[-/](0[1-9]|1[0-2])\b", qn):
        explicit.append(f"{y}-{m}")
    if explicit:
        return [m for m in sorted(set(explicit)) if m in months]
    current = txt(filters.get("current_period"))
    if current and current in months:
        return [current]
    return [months[-1]] if months else []


def wants_compare(qn: str) -> bool:
    return any(term in qn for term in [" vs ", "ly", "año pasado", "ano pasado", "compar", "delta", "contra"])


def wants_chart(qn: str) -> bool:
    return any(term in qn for term in ["graf", "chart", "barra", "barras", "pastel", "pie", "visual", "muestrame"]) 


def chart_type(qn: str) -> str:
    if any(term in qn for term in ["pastel", "pie", "donut", "doughnut"]):
        return "pie"
    if any(term in qn for term in ["linea", "línea"]):
        return "line"
    return "bar"


def top_n(qn: str) -> int:
    m = re.search(r"\btop\s*(\d{1,2})\b", qn)
    if m:
        return max(1, min(20, int(m.group(1))))
    return 10


def metric_for(domain: str, qn: str) -> str:
    if domain in {"sales", "delivery", "innovation"}:
        if any(term in qn for term in ["cantidad", "qty", "unidades"]):
            return "Qty Sold"
        if "margen" in qn:
            return "Total Gross Margin"
        return "Sales"
    if domain == "sbf":
        return "TX"
    if domain == "inventory":
        if any(term in qn for term in ["dias inv", "doi"]):
            return "Días Inv"
        if any(term in qn for term in ["cantidad", "qty"]):
            return "Quantity"
        if "sales s4" in qn:
            return "Sales S4"
        return "Amount"
    if domain == "merma":
        return "Merma"
    if domain == "cxc":
        return "Amount"
    if domain == "prd0":
        return "Lost Sales 3d"
    return "Sales"


def pick_group(qn: str, domain: str) -> Optional[str]:
    for alias, col in GROUP_ALIASES.items():
        if alias in qn:
            if domain == "sbf" and col not in {"Store", "Provider", "Daypart", "DOW_Name", "Hour"}:
                continue
            if domain in {"inventory", "merma", "cxc", "prd0"} and col in {"Date Sold", "Delivery_Channel", "ComboLabel", "Hour"}:
                continue
            return col
    return None


def domain_scores(qn: str, active_tab: str) -> Dict[str, int]:
    scores = {k: 0 for k in ["sales", "delivery", "sbf", "inventory", "merma", "innovation", "cxc", "prd0", "hallazgos"]}
    def add(domain: str, points: int, *terms: str) -> None:
        for term in terms:
            if term in qn:
                scores[domain] += points
    add("delivery", 7, "delivery", "pedido ya", "pedidos ya", "uber", "canal")
    add("sbf", 8, "sbf", "aki pago", "puntoxpress", "western union", "servicio")
    add("inventory", 8, "inventario", "stock", "doi", "dias inv", "planimetria", "transfer")
    add("merma", 8, "merma", "ajuste", "desperdicio")
    add("innovation", 8, "innovacion", "food service", "combo", "hot dog", "pizza", "hamburg", "budin")
    add("cxc", 8, "cxc", "faltante", "sobrante", "caja", "batch")
    add("prd0", 8, "prd en cero", "prd0", "quiebre", "agotado", "lost sales")
    add("hallazgos", 7, "hallazgo", "hallazgos", "lectura ejecutiva", "riesgo", "oportunidad", "acciones")
    add("sales", 3, "ventas", "venta", "sales", "categoria", "category", "marca", "brand", "descripcion", "description")
    if active_tab in scores:
        scores[active_tab] += 2
    return scores


def pick_domain(qn: str, active_tab: str) -> Tuple[str, List[str]]:
    ranked = sorted(domain_scores(qn, active_tab).items(), key=lambda kv: kv[1], reverse=True)
    explicit = [name for name, score in ranked if score >= 7]
    if len(explicit) >= 2:
        return explicit[0], explicit[:2]
    if ranked[0][1] > 0:
        return ranked[0][0], []
    return (active_tab if active_tab in DOMAIN_LABELS else "sales"), []


def apply_named_filters(df: pd.DataFrame, qn: str, domain: str) -> pd.DataFrame:
    out = df.copy()
    def filter_contains(cols: List[str], needle: str) -> pd.DataFrame:
        tmp = out
        for col in cols:
            ncol = f"_{col}"
            if ncol in tmp.columns:
                mask = tmp[ncol].str.contains(needle, na=False)
                if mask.any():
                    return tmp[mask].copy()
        return tmp
    if domain in {"sales", "delivery", "innovation", "inventory", "merma", "prd0"}:
        if any(t in qn for t in ["cigarro", "cigarrillo", "cigarrillos"]):
            out = filter_contains(["Category", "Department", "Description", "Brand"], "cigarr")
        elif any(t in qn for t in ["vape", "vapes"]):
            out = filter_contains(["Category", "Department", "Description", "Brand"], "vape")
        elif any(t in qn for t in ["licor", "licores"]):
            out = filter_contains(["Category", "Department", "Description", "Brand"], "licor")
    if domain == "delivery" and "Delivery_Channel" in out.columns:
        if any(t in qn for t in ["pedido ya", "pedidos ya", "pedidoya"]):
            mask = out["_Delivery_Channel"].str.contains("pedido", na=False)
            if mask.any(): out = out[mask].copy()
        elif "uber" in qn:
            mask = out["_Delivery_Channel"].str.contains("uber", na=False)
            if mask.any(): out = out[mask].copy()
    if domain == "sbf" and "Provider" in out.columns:
        for term in ["aki pago", "puntoxpress", "western union", "smart", "bac", "cuscatlan"]:
            if term in qn:
                mask = out["_Provider"].str.contains(term.split()[0], na=False)
                if mask.any():
                    out = out[mask].copy()
                    break
    return out


def default_actions(domain: str) -> List[str]:
    if domain == "sales":
        return ["Ataca primero la línea líder: ahí está la mayor palanca comercial.", "Si el delta cae, revisa disponibilidad, precio y ejecución antes de abrir campaña.", "Replica en otras tiendas la categoría o marca que ya empuja el resultado."]
    if domain == "delivery":
        return ["Concentra ejecución en el canal y franja que realmente aportan venta.", "Si LY cae, revisa disponibilidad y tiempos por tienda antes de empujar pauta.", "Usa la tienda top como benchmark operativo del canal."]
    if domain == "sbf":
        return ["Protege cobertura en el servicio y tienda con mayor peso transaccional.", "Si el delta cae, revisa visibilidad del servicio y disciplina de caja.", "Replica horarios ganadores donde la TX ya demuestra tracción."]
    if domain == "inventory":
        return ["Prioriza exceso y días de inventario altos en las líneas de menor rotación.", "Evita recompras donde la cobertura ya está sobrada.", "Protege disponibilidad en SKU de alta venta S4 antes de mover capital al resto."]
    if domain == "merma":
        return ["Ataca primero las descripciones con mayor impacto económico absoluto.", "Cruza merma con venta para separar ruido de problema material.", "Corrige causa raíz en tienda antes de ampliar pedido del mismo SKU."]
    if domain == "innovation":
        return ["Escala el combo o categoría ganadora donde ya muestra tracción.", "Si un bloque no rota, corrige surtido o comunicación antes de ampliar portafolio.", "Usa el mix top como base del siguiente test comercial."]
    if domain == "cxc":
        return ["Ataca primero el batch o comentario con mayor monto absoluto.", "Separa eventos compensados de faltantes reales antes de decidir.", "Refuerza disciplina de caja en la tienda con mayor recurrencia."]
    if domain == "prd0":
        return ["Prioriza los SKU con mayor venta perdida estimada en 3 días.", "Corrige ROP y compra en líneas de alta rotación.", "Usa reposición diaria en los quiebres realmente materiales."]
    return ["Convierte el hallazgo en una acción operativa y una acción comercial."]


def build_chart(labels: List[str], datasets: List[Dict[str, Any]], title: str, subtitle: str, qn: str) -> Dict[str, Any]:
    return {"title": title, "subtitle": subtitle, "type": chart_type(qn), "labels": labels, "datasets": datasets, "note": "Fuente: Cubo_Semanal_Compactado.xlsx"}


def summarize_preview(grouped: pd.DataFrame, group_col: str, metric_col: str, limit: int = 5) -> str:
    items = []
    for _, row in grouped.head(limit).iterrows():
        label = txt(row[group_col]) or "Sin etiqueta"
        val = float(row[metric_col] or 0)
        rendered = fmt_num(val) if metric_col in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv"} else fmt_money(val)
        items.append(f"{label}: {rendered}")
    return " | ".join(items)


def resolve_general(cache: SourceCache, domain: str, question: str, filters: Dict[str, Any], page_context: Dict[str, Any]) -> Dict[str, Any]:
    qn = norm(question)
    sales_df = get_frame(cache, "sales")
    stores = extract_stores(qn, filters, sales_df)

    frame_key = {"sales": "sales", "delivery": "sales", "sbf": "sbf_month", "inventory": "inventory", "merma": "merma", "innovation": "innovation", "cxc": "cxc", "prd0": "prd0"}[domain]
    df = get_frame(cache, frame_key)
    months = available_months(df if frame_key != "sales" else sales_df)
    chosen_months = extract_months(qn, filters, months) if months else []

    if domain == "sales":
        df = df[df["Is_Delivery"] == False].copy()  # noqa: E712
    elif domain == "delivery":
        df = df[df["Is_Delivery"] == True].copy()  # noqa: E712
    if stores and "Store" in df.columns:
        df = df[df["Store"].isin(stores)].copy()
    if chosen_months:
        period_col = "CalMonth" if "CalMonth" in df.columns else "Month" if "Month" in df.columns else None
        if period_col:
            df = df[df[period_col].isin(chosen_months)].copy()
    if filters.get("day_of_week") and "DOW" in df.columns:
        vals = []
        for raw in filters.get("day_of_week") or []:
            try: vals.append(int(raw))
            except Exception: pass
        if vals:
            df = df[df["DOW"].isin(vals)].copy()
    df = apply_named_filters(df, qn, domain)

    if domain in {"sales", "delivery", "innovation"} and "hora" in qn:
        return {"title": f"{DOMAIN_LABELS[domain]} · hora no disponible", "summary": "La fuente actual no trae hora real para este dominio. Sí trae día de semana y horario operativo, así que la consulta correcta es por día de semana o por daypart.", "actions": ["Pídelo por día de semana para tener una lectura utilizable ya.", "Pídelo por horario operativo para capturar la señal comercial disponible.", "Si quieres hora real, agrega Hour a la fuente de este dominio."], "notes": [f"Fuente de verdad del chat: {cache.source}"], "chart": None}
    if df.empty:
        return {"title": f"{DOMAIN_LABELS[domain]} · sin data visible", "summary": "No encontré filas visibles con la combinación actual de periodo, tienda y filtros de la pregunta.", "actions": ["Revisa si la tienda o el mes pedido existe en la fuente publicada.", "Prueba la misma pregunta sin filtro de tienda para validar cobertura.", "Si esperabas un grano más fino, revisa si esa dimensión existe en el workbook."], "notes": [f"Fuente de verdad del chat: {cache.source}"], "chart": None}

    metric = metric_for(domain, qn)
    group_col = pick_group(qn, domain)
    compare = wants_compare(qn)
    n = top_n(qn)
    current_label = ", ".join(month_label(m) for m in chosen_months) if chosen_months else "corte actual"

    if group_col is None and (wants_chart(qn) or "top" in qn):
        defaults = {"sales": "Category", "delivery": "Store", "sbf": "Store", "inventory": "Store", "merma": "Description", "innovation": "ComboLabel", "cxc": "Store", "prd0": "Description"}
        group_col = defaults.get(domain)

    if compare and len(chosen_months) == 1 and frame_key in {"sales", "sbf_month", "merma", "innovation"}:
        ly = same_month_ly(chosen_months[0])
        if ly:
            base_all = get_frame(cache, frame_key)
            if domain == "sales":
                base_all = base_all[base_all["Is_Delivery"] == False].copy()  # noqa: E712
            elif domain == "delivery":
                base_all = base_all[base_all["Is_Delivery"] == True].copy()  # noqa: E712
            if stores and "Store" in base_all.columns:
                base_all = base_all[base_all["Store"].isin(stores)].copy()
            period_col = "CalMonth" if "CalMonth" in base_all.columns else "Month"
            prev = base_all[base_all[period_col] == ly].copy()
            prev = apply_named_filters(prev, qn, domain)
            if group_col and group_col in df.columns and group_col in prev.columns:
                curg = df.groupby(group_col, dropna=False)[metric].sum().reset_index().rename(columns={metric: "current"})
                prevg = prev.groupby(group_col, dropna=False)[metric].sum().reset_index().rename(columns={metric: "ly"})
                merged = curg.merge(prevg, on=group_col, how="outer").fillna(0)
                merged[group_col] = merged[group_col].fillna("").astype(str)
                merged = merged[merged[group_col].str.strip() != ""]
                merged = merged.sort_values("current", ascending=False).head(n)
                cur_total = float(merged["current"].sum())
                ly_total = float(merged["ly"].sum())
                chart = build_chart([str(v) for v in merged[group_col].tolist()], [{"label": month_label(chosen_months[0]), "data": [float(v) for v in merged['current'].tolist()]}, {"label": month_label(ly), "data": [float(v) for v in merged['ly'].tolist()]}], f"{DOMAIN_LABELS[domain]} por {group_col}", f"{month_label(chosen_months[0])} vs LY", qn)
                return {"title": f"{DOMAIN_LABELS[domain]} por {group_col} · vs LY", "summary": f"Periodo {month_label(chosen_months[0])}. Visible actual: {fmt_num(cur_total) if metric=='TX' else fmt_money(cur_total)} vs LY {fmt_num(ly_total) if metric=='TX' else fmt_money(ly_total)}. Delta: {fmt_pct(pct_delta(cur_total, ly_total))}. Top visible: {summarize_preview(merged.rename(columns={'current': metric}), group_col, metric)}.", "actions": default_actions(domain), "notes": [f"Tiendas: {', '.join(stores) if stores else 'todas'}", f"Fuente de verdad del chat: {cache.source}"], "chart": chart}
            else:
                cur_total = float(df[metric].sum())
                ly_total = float(prev[metric].sum()) if not prev.empty else 0.0
                chart = build_chart([month_label(chosen_months[0]), month_label(ly)], [{"label": DOMAIN_LABELS[domain], "data": [cur_total, ly_total]}], f"{DOMAIN_LABELS[domain]} total", f"{month_label(chosen_months[0])} vs LY", qn)
                return {"title": f"{DOMAIN_LABELS[domain]} total · vs LY", "summary": f"Periodo {month_label(chosen_months[0])}. Actual: {fmt_num(cur_total) if metric=='TX' else fmt_money(cur_total)} vs LY {fmt_num(ly_total) if metric=='TX' else fmt_money(ly_total)}. Delta: {fmt_pct(pct_delta(cur_total, ly_total))}.", "actions": default_actions(domain), "notes": [f"Tiendas: {', '.join(stores) if stores else 'todas'}", f"Fuente de verdad del chat: {cache.source}"], "chart": chart}

    if group_col and group_col in df.columns:
        grouped = df.groupby(group_col, dropna=False)[metric].sum().reset_index()
        grouped[group_col] = grouped[group_col].fillna("").astype(str)
        grouped = grouped[grouped[group_col].str.strip() != ""]
        grouped = grouped.sort_values(metric, ascending=False).head(n)
        total = float(df[metric].sum())
        chart = build_chart([str(v) for v in grouped[group_col].tolist()], [{"label": metric, "data": [float(v) for v in grouped[metric].tolist()]}], f"{DOMAIN_LABELS[domain]} por {group_col}", f"Corte {current_label}", qn) if wants_chart(qn) or "top" in qn else None
        rendered_total = fmt_num(total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv"} else fmt_money(total)
        return {"title": f"{DOMAIN_LABELS[domain]} por {group_col}", "summary": f"Corte {current_label}. Total visible: {rendered_total}. Top visible: {summarize_preview(grouped, group_col, metric)}.", "actions": default_actions(domain), "notes": [f"Tiendas: {', '.join(stores) if stores else 'todas'}", f"Fuente de verdad del chat: {cache.source}"], "chart": chart}

    total = float(df[metric].sum())
    rendered_total = fmt_num(total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv"} else fmt_money(total)
    return {"title": f"{DOMAIN_LABELS[domain]} total", "summary": f"Corte {current_label}. Total visible: {rendered_total}. Consulta resuelta sobre la fuente canónica publicada en GitHub.", "actions": default_actions(domain), "notes": [f"Tiendas: {', '.join(stores) if stores else 'todas'}", f"Fuente de verdad del chat: {cache.source}"], "chart": None}


def resolve_hallazgos(cache: SourceCache, question: str, filters: Dict[str, Any]) -> Dict[str, Any]:
    sales = get_frame(cache, "sales")
    stores = extract_stores(norm(question), filters, sales)
    if stores:
        sales = sales[sales["Store"].isin(stores)].copy()
    months = available_months(sales)
    month_key = extract_months(norm(question), filters, months)[0] if months else ""
    ly = same_month_ly(month_key) if month_key else None
    cur_sales = sales[sales["CalMonth"] == month_key]
    prev_sales = sales[sales["CalMonth"] == ly] if ly else sales.iloc[0:0]
    cur_total = float(cur_sales["Sales"].sum())
    prev_total = float(prev_sales["Sales"].sum()) if not prev_sales.empty else 0.0
    delivery = sales[sales["Is_Delivery"] == True]
    cur_del = delivery[delivery["CalMonth"] == month_key]
    prev_del = delivery[delivery["CalMonth"] == ly] if ly else delivery.iloc[0:0]
    cur_sbf_df = get_frame(cache, "sbf_month")
    if stores:
        cur_sbf_df = cur_sbf_df[cur_sbf_df["Store"].isin(stores)].copy()
    cur_sbf = cur_sbf_df[cur_sbf_df["CalMonth"] == month_key]
    prev_sbf = cur_sbf_df[cur_sbf_df["CalMonth"] == ly] if ly else cur_sbf_df.iloc[0:0]
    merma_df = get_frame(cache, "merma")
    if stores:
        merma_df = merma_df[merma_df["Store"].isin(stores)].copy()
    cur_merma = merma_df[merma_df["CalMonth"] == month_key]
    top_merma = cur_merma.groupby("Description", dropna=False)["Merma"].sum().reset_index().sort_values("Merma", ascending=True).head(3)
    summary = "\n\n".join([
        f"Ventas {month_label(month_key)}: {fmt_money(cur_total)} vs LY {fmt_money(prev_total)} ⇒ {fmt_pct(pct_delta(cur_total, prev_total))}.",
        f"Delivery {month_label(month_key)}: {fmt_money(float(cur_del['Sales'].sum()))} vs LY {fmt_money(float(prev_del['Sales'].sum()) if not prev_del.empty else 0.0)} ⇒ {fmt_pct(pct_delta(float(cur_del['Sales'].sum()), float(prev_del['Sales'].sum()) if not prev_del.empty else 0.0))}.",
        f"SBF {month_label(month_key)}: {fmt_num(float(cur_sbf['TX'].sum()))} TX vs LY {fmt_num(float(prev_sbf['TX'].sum()) if not prev_sbf.empty else 0.0)} ⇒ {fmt_pct(pct_delta(float(cur_sbf['TX'].sum()), float(prev_sbf['TX'].sum()) if not prev_sbf.empty else 0.0))}.",
        (f"Merma visible {month_label(month_key)}: {fmt_money(float(cur_merma['Merma'].sum()))}. Drivers: " + " | ".join([f"{txt(r['Description'])}: {fmt_money(float(r['Merma']))}" for _, r in top_merma.iterrows()])) if not top_merma.empty else f"Merma visible {month_label(month_key)}: {fmt_money(float(cur_merma['Merma'].sum()))}.",
    ])
    return {"title": "Hallazgos ejecutivos", "summary": summary, "actions": ["Ataca primero la señal con peor delta absoluto y dueño claro por tienda.", "Convierte el driver principal en una acción operativa y otra comercial.", "Haz seguimiento semanal del mismo corte para validar impacto real."], "notes": [f"Tiendas: {', '.join(stores) if stores else 'todas'}", f"Fuente de verdad del chat: {cache.source}"], "chart": None}


def resolve_query(cache: SourceCache, payload: QueryRequest) -> Dict[str, Any]:
    question = txt(payload.question)
    if not question:
        raise HTTPException(status_code=400, detail="La pregunta viene vacía.")
    qn = norm(question)
    active_tab = norm(payload.page_context.get("tab"))
    domain, conflict = pick_domain(qn, active_tab)
    if conflict:
        return {"title": "Consulta ambigua", "summary": f"La pregunta mezcla dos dominios distintos: {DOMAIN_LABELS.get(conflict[0], conflict[0])} y {DOMAIN_LABELS.get(conflict[1], conflict[1])}. Para responder bien, sepárala en dos consultas.", "actions": [f"Ejemplo 1: dame {DOMAIN_LABELS.get(conflict[0], conflict[0]).lower()} por tienda con delta vs LY.", f"Ejemplo 2: dame {DOMAIN_LABELS.get(conflict[1], conflict[1]).lower()} por tienda con delta vs LY.", "No mezcles dos motores de negocio en una misma pregunta si quieres cifras confiables."], "notes": [f"Fuente de verdad del chat: {cache.source}"], "chart": None}
    if domain == "hallazgos" or any(t in qn for t in ["hallazgo", "hallazgos", "lectura ejecutiva", "riesgo", "oportunidad"]):
        return resolve_hallazgos(cache, question, payload.filters)
    return resolve_general(cache, domain, question, payload.filters, payload.page_context)


@app.get('/health')
def health() -> Dict[str, Any]:
    return {"ok": True, "version": APP_VERSION, "cache_ttl_seconds": CACHE_TTL_SECONDS, "default_data_url": DEFAULT_DATA_URL or None, "allowed_origins": allow_origins}


@app.post('/api/query')
def query(payload: QueryRequest) -> Dict[str, Any]:
    source = txt(payload.data_url) or DEFAULT_DATA_URL
    if not source:
        raise HTTPException(status_code=400, detail="No llegó data_url y WORKBOOK_PUBLIC_URL no está configurado.")
    try:
        cache = get_source_cache(source)
        result = resolve_query(cache, payload)
        return {"ok": True, "engine": APP_VERSION, "source": source, "result": result}
    except HTTPException:
        raise
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"No pude descargar la fuente publicada: {exc}") from exc
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"No encontré el workbook: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error resolviendo la consulta: {exc}") from exc
