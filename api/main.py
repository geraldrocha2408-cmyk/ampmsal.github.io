import os
import re
import time
import unicodedata
from io import BytesIO

import httpx
import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook

app = FastAPI(title="AMPM Data API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CACHE = {
    "ts": 0,
    "excel_bytes": None,
    "sheet_names": None,
    "sales_pack": None,
    "tx_pack": None,
}

CACHE_SECONDS = 300


def strip_accents(value: str) -> str:
    value = str(value or "").strip()
    return "".join(
        c for c in unicodedata.normalize("NFD", value)
        if unicodedata.category(c) != "Mn"
    )


def norm(value: str) -> str:
    return strip_accents(value).lower().strip()


def pick_col(columns, aliases):
    norm_map = {norm(col): col for col in columns}

    for alias in aliases:
        key = norm(alias)
        if key in norm_map:
            return norm_map[key]

    for alias in aliases:
        key = norm(alias)
        for ncol, original in norm_map.items():
            if key in ncol or ncol in key:
                return original

    return None


def parse_month_key(value):
    if pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        return f"{value.year:04d}-{value.month:02d}"

    s = str(value).strip()
    ns = norm(s)

    if "sem" in ns or "week" in ns or re.search(r"\bs\d{1,2}\b", ns):
        return None

    m = re.match(r"^(\d{4})[-/_ ](\d{1,2})$", ns)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}"

    m = re.match(r"^(\d{4})(\d{2})$", ns)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}"

    m = re.match(r"^(\d{4})m(\d{1,2})$", ns)
    if m:
        y, mo = int(m.group(1)), int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}-{mo:02d}"

    try:
        dt = pd.to_datetime(value, errors="raise")
        return f"{dt.year:04d}-{dt.month:02d}"
    except Exception:
        return None


async def ensure_excel_bytes():
    excel_url = os.getenv("EXCEL_URL", "").strip()
    if not excel_url:
        raise HTTPException(status_code=500, detail="Falta la variable EXCEL_URL")

    now = time.time()
    if CACHE["excel_bytes"] is not None and (now - CACHE["ts"]) < CACHE_SECONDS:
        return CACHE["excel_bytes"]

    async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
        response = await client.get(excel_url)
        response.raise_for_status()
        content = response.content

    CACHE["ts"] = now
    CACHE["excel_bytes"] = content
    CACHE["sheet_names"] = None
    CACHE["sales_pack"] = None
    CACHE["tx_pack"] = None
    return content


async def get_sheet_names():
    if CACHE["sheet_names"] is not None:
        return CACHE["sheet_names"]

    content = await ensure_excel_bytes()
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    CACHE["sheet_names"] = sheet_names
    return sheet_names


async def get_sales_pack():
    if CACHE["sales_pack"] is not None:
        return CACHE["sales_pack"]

    content = await ensure_excel_bytes()
    sheet_names = await get_sheet_names()

    sales_sheet = None
    for name in ["Cubo_Sales_DetailHallazgos", "Cubo_Sales", "Cubo_Sales_Fact"]:
        if name in sheet_names:
            sales_sheet = name
            break

    if not sales_sheet:
        raise HTTPException(status_code=500, detail="No encontré hoja de ventas")

    df = pd.read_excel(BytesIO(content), sheet_name=sales_sheet)
    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)

    c_store = pick_col(cols, ["Store", "Store Name", "Tienda"])
    c_period = pick_col(cols, ["YearMonth", "YM", "Year Month", "Month", "Mes", "CalMonth", "CalendarMonth", "Fecha", "Date"])
    c_dept = pick_col(cols, ["Department", "Departamento"])
    c_cat = pick_col(cols, ["Category", "Categoria", "Categoría"])
    c_brand = pick_col(cols, ["Brand", "Marca"])
    c_sales = pick_col(cols, ["Sales", "Total Sales", "TotalSales", "Venta", "Ventas"])
    c_margin = pick_col(cols, ["Total Gross Margin", "Gross Margin", "Margen", "GM"])

    if not c_period or not c_sales:
        raise HTTPException(status_code=500, detail=f"No pude identificar columnas base en ventas. Columnas detectadas: {cols}")

    out = df.copy()
    out["_month_key"] = out[c_period].apply(parse_month_key)
    out = out[out["_month_key"].notna()].copy()

    out["_store"] = out[c_store].astype(str).str.strip().str.upper() if c_store else "TOTAL"
    out["_department"] = out[c_dept].astype(str).str.strip() if c_dept else ""
    out["_category"] = out[c_cat].astype(str).str.strip() if c_cat else ""
    out["_brand"] = out[c_brand].astype(str).str.strip() if c_brand else ""
    out["_sales"] = pd.to_numeric(out[c_sales], errors="coerce").fillna(0)
    out["_margin"] = pd.to_numeric(out[c_margin], errors="coerce").fillna(0) if c_margin else 0

    CACHE["sales_pack"] = {
        "sheet": sales_sheet,
        "columns": cols,
        "df": out
    }
    return CACHE["sales_pack"]


async def get_tx_pack():
    if CACHE["tx_pack"] is not None:
        return CACHE["tx_pack"]

    content = await ensure_excel_bytes()
    sheet_names = await get_sheet_names()

    tx_sheet = None
    for name in ["Cubo_TX_Fact", "Cubo_TX", "Cubo_TX_Month_Fact"]:
        if name in sheet_names:
            tx_sheet = name
            break

    if not tx_sheet:
        CACHE["tx_pack"] = {"sheet": None, "columns": [], "df": None}
        return CACHE["tx_pack"]

    df = pd.read_excel(BytesIO(content), sheet_name=tx_sheet)
    df.columns = [str(c).strip() for c in df.columns]
    cols = list(df.columns)

    c_store = pick_col(cols, ["Store", "Store Name", "Tienda"])
    c_period = pick_col(cols, ["YearMonth", "YM", "Year Month", "Month", "Mes", "CalMonth", "CalendarMonth", "Fecha", "Date"])
    c_tx = pick_col(cols, ["Transactions", "TX", "Transacciones"])

    if not c_period or not c_tx:
        CACHE["tx_pack"] = {"sheet": tx_sheet, "columns": cols, "df": None}
        return CACHE["tx_pack"]

    out = df.copy()
    out["_month_key"] = out[c_period].apply(parse_month_key)
    out = out[out["_month_key"].notna()].copy()

    out["_store"] = out[c_store].astype(str).str.strip().str.upper() if c_store else "TOTAL"
    out["_tx"] = pd.to_numeric(out[c_tx], errors="coerce").fillna(0)

    CACHE["tx_pack"] = {
        "sheet": tx_sheet,
        "columns": cols,
        "df": out
    }
    return CACHE["tx_pack"]


@app.get("/health")
async def health():
    return {"ok": True}


@app.get("/meta")
async def meta():
    sheet_names = await get_sheet_names()
    sales_pack = await get_sales_pack()
    tx_pack = await get_tx_pack()

    months = sorted(sales_pack["df"]["_month_key"].dropna().unique().tolist())

    return {
        "ok": True,
        "sheets": sheet_names,
        "sales_sheet": sales_pack["sheet"],
        "sales_columns": sales_pack["columns"],
        "tx_sheet": tx_pack["sheet"],
        "tx_columns": tx_pack["columns"],
        "sample_months": months[:36]
    }


@app.get("/sheet-preview")
async def sheet_preview(
    sheet: str = Query(...),
    rows: int = Query(5, ge=1, le=20)
):
    content = await ensure_excel_bytes()
    df = pd.read_excel(BytesIO(content), sheet_name=sheet, nrows=rows)
    df.columns = [str(c).strip() for c in df.columns]

    return {
        "ok": True,
        "sheet": sheet,
        "columns": df.columns.tolist(),
        "preview_rows": df.fillna("").to_dict(orient="records")
    }


@app.get("/monthly-category-sales")
async def monthly_category_sales(
    year: int = Query(...),
    month: int = Query(...),
    department: str = Query(...),
    store: str = Query(None)
):
    sales_pack = await get_sales_pack()
    df = sales_pack["df"].copy()
    month_key = f"{year:04d}-{month:02d}"

    df = df[df["_month_key"] == month_key]
    df = df[df["_department"].apply(norm) == norm(department)]

    if store:
        df = df[df["_store"] == store.strip().upper()]

    if df.empty:
        return {
            "ok": True,
            "month_key": month_key,
            "department": department,
            "store": store,
            "labels": [],
            "values": [],
            "total": 0
        }

    grouped = (
        df.groupby("_category", dropna=False)["_sales"]
        .sum()
        .reset_index()
        .sort_values("_sales", ascending=False)
    )

    grouped["_category"] = grouped["_category"].replace("", "Sin categoría")

    return {
        "ok": True,
        "month_key": month_key,
        "department": department,
        "store": store,
        "labels": grouped["_category"].astype(str).tolist(),
        "values": grouped["_sales"].round(2).tolist(),
        "total": round(float(grouped["_sales"].sum()), 2)
    }


@app.get("/monthly-compare")
async def monthly_compare(
    year: int = Query(...),
    month: int = Query(...),
    compare_year: int = Query(...),
    store: str = Query(None)
):
    sales_pack = await get_sales_pack()
    tx_pack = await get_tx_pack()

    sales_df = sales_pack["df"].copy()
    tx_df = tx_pack["df"].copy() if tx_pack["df"] is not None else None

    cur_key = f"{year:04d}-{month:02d}"
    ly_key = f"{compare_year:04d}-{month:02d}"

    cur_sales = sales_df[sales_df["_month_key"] == cur_key]
    ly_sales = sales_df[sales_df["_month_key"] == ly_key]

    if store:
        store_norm = store.strip().upper()
        cur_sales = cur_sales[cur_sales["_store"] == store_norm]
        ly_sales = ly_sales[ly_sales["_store"] == store_norm]

    sales = float(cur_sales["_sales"].sum()) if not cur_sales.empty else 0
    sales_ly = float(ly_sales["_sales"].sum()) if not ly_sales.empty else 0

    margin = float(cur_sales["_margin"].sum()) if not cur_sales.empty else 0
    margin_ly = float(ly_sales["_margin"].sum()) if not ly_sales.empty else 0

    tx = None
    tx_ly = None

    if tx_df is not None:
        cur_tx = tx_df[tx_df["_month_key"] == cur_key]
        ly_tx = tx_df[tx_df["_month_key"] == ly_key]

        if store:
            store_norm = store.strip().upper()
            cur_tx = cur_tx[cur_tx["_store"] == store_norm]
            ly_tx = ly_tx[ly_tx["_store"] == store_norm]

        tx = float(cur_tx["_tx"].sum()) if not cur_tx.empty else 0
        tx_ly = float(ly_tx["_tx"].sum()) if not ly_tx.empty else 0

    avg_ticket = (sales / tx) if tx not in (None, 0) else None
    avg_ticket_ly = (sales_ly / tx_ly) if tx_ly not in (None, 0) else None
    margin_pct = (margin / sales * 100) if sales else None
    margin_pct_ly = (margin_ly / sales_ly * 100) if sales_ly else None

    return {
        "ok": True,
        "current_period": cur_key,
        "compare_period": ly_key,
        "store": store,
        "sales": round(sales, 2),
        "sales_ly": round(sales_ly, 2),
        "sales_delta": round(sales - sales_ly, 2),
        "sales_delta_pct": round(((sales / sales_ly) - 1) * 100, 2) if sales_ly else None,
        "tx": round(tx, 2) if tx is not None else None,
        "tx_ly": round(tx_ly, 2) if tx_ly is not None else None,
        "tx_delta": round(tx - tx_ly, 2) if tx is not None and tx_ly is not None else None,
        "tx_delta_pct": round(((tx / tx_ly) - 1) * 100, 2) if tx not in (None, 0) and tx_ly not in (None, 0) else None,
        "avg_ticket": round(avg_ticket, 2) if avg_ticket is not None else None,
        "avg_ticket_ly": round(avg_ticket_ly, 2) if avg_ticket_ly is not None else None,
        "avg_ticket_delta_pct": round(((avg_ticket / avg_ticket_ly) - 1) * 100, 2) if avg_ticket is not None and avg_ticket_ly not in (None, 0) else None,
        "margin_pct": round(margin_pct, 2) if margin_pct is not None else None,
        "margin_pct_ly": round(margin_pct_ly, 2) if margin_pct_ly is not None else None
    }