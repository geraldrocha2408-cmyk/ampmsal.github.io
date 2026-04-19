from __future__ import annotations

import io
import math
import os
import re
import threading
import time
import unicodedata
from dataclasses import dataclass, field
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import requests
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field

APP_VERSION = "ampm-chat-engine-v3"
CACHE_TTL_SECONDS = int(os.getenv("WORKBOOK_CACHE_TTL_SECONDS", "900"))
DEFAULT_DATA_URL = os.getenv("WORKBOOK_PUBLIC_URL", "").strip()
REQUEST_TIMEOUT = int(os.getenv("WORKBOOK_REQUEST_TIMEOUT_SECONDS", "90"))

MONTHS_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}
MONTHS_ABBR = {1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun", 7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"}
DOW_NUM_TO_ES = {1: "Lunes", 2: "Martes", 3: "Miércoles", 4: "Jueves", 5: "Viernes", 6: "Sábado", 7: "Domingo"}

SHEET_CONFIG: Dict[str, Dict[str, Any]] = {
    "sales": {
        "sheet": "Cubo_Sales_DetailHallazgos",
        "columns": [
            "Date Sold", "YearWeek", "CalMonth", "Store", "Department", "Category", "Supplier", "Brand", "Description",
            "Daypart", "DOW", "DOW_Name", "Is_Delivery", "Delivery_Channel", "Qty Sold", "Sales", "Total Gross Margin",
        ],
        "text": ["Store", "Department", "Category", "Supplier", "Brand", "Description", "Daypart", "DOW_Name", "Delivery_Channel"],
        "numeric": ["Qty Sold", "Sales", "Total Gross Margin", "DOW"],
        "dates": ["Date Sold"],
    },
    "sbf": {
        "sheet": "Chart_SBF_Month_XF",
        "columns": ["CalMonth", "Store", "Provider", "Daypart", "DOW", "Hour", "TX"],
        "text": ["Store", "Provider", "Daypart"],
        "numeric": ["TX", "DOW", "Hour"],
        "dates": [],
    },
    "inventory": {
        "sheet": "Inventario_Detail",
        "columns": ["Store", "Department", "Supplier", "Category", "Brand", "Description", "Amount", "Quantity", "Días Inv", "Qty Sold S4", "Sales S4", "IdealQty", "TargetQty"],
        "text": ["Store", "Department", "Supplier", "Category", "Brand", "Description"],
        "numeric": ["Amount", "Quantity", "Días Inv", "Qty Sold S4", "Sales S4", "IdealQty", "TargetQty"],
        "dates": [],
    },
    "merma": {
        "sheet": "Merma_Detail",
        "columns": ["Store", "CalMonth", "Supplier", "Category", "Brand", "Description", "Total Sales", "Merma", "Quantity (packs)"],
        "text": ["Store", "Supplier", "Category", "Brand", "Description"],
        "numeric": ["Total Sales", "Merma", "Quantity (packs)"],
        "dates": [],
    },
    "innovation": {
        "sheet": "Innovation_Combos_Detail",
        "columns": [
            "Date Sold", "YearWeek", "CalMonth", "Store", "Department", "Category", "Supplier", "Brand", "Description",
            "ComboLabel", "Daypart", "DOW", "DOW_Name", "Is_Delivery", "Delivery_Channel", "Qty Sold", "Sales", "Total Gross Margin",
        ],
        "text": ["Store", "Department", "Category", "Supplier", "Brand", "Description", "ComboLabel", "Daypart", "DOW_Name", "Delivery_Channel"],
        "numeric": ["Qty Sold", "Sales", "Total Gross Margin", "DOW"],
        "dates": ["Date Sold"],
    },
    "cxc": {
        "sheet": "CXC_DETAIL",
        "columns": ["Store", "BatchNumber", "Time", "Month", "Comment", "Cashier", "Amount"],
        "text": ["Store", "BatchNumber", "Comment", "Cashier"],
        "numeric": ["Amount"],
        "dates": ["Time"],
    },
    "prd0": {
        "sheet": "PRD_CERO_DETAIL",
        "columns": ["Store", "Supplier", "Description", "Department", "Category", "Brand", "Quantity", "ROP", "Qty Sold S4", "Sales S4", "Lost Units 3d", "Lost Sales 3d"],
        "text": ["Store", "Supplier", "Description", "Department", "Category", "Brand"],
        "numeric": ["Quantity", "ROP", "Qty Sold S4", "Sales S4", "Lost Units 3d", "Lost Sales 3d"],
        "dates": [],
    },
}

DOMAIN_LABELS = {
    "sales": "Ventas",
    "delivery": "Delivery",
    "sbf": "SBF",
    "inventory": "Inventario",
    "merma": "Merma",
    "innovation": "Innovación",
    "cxc": "CXC",
    "prd0": "PRD en cero",
    "hallazgos": "Hallazgos",
}

class QueryRequest(BaseModel):
    question: str
    filters: Dict[str, Any] = Field(default_factory=dict)
    page_context: Dict[str, Any] = Field(default_factory=dict)
    data_url: Optional[str] = None

@dataclass
class ValueIndex:
    values_by_col: Dict[str, List[Tuple[str, List[str]]]] = field(default_factory=dict)
    months: List[str] = field(default_factory=list)
    stores: List[str] = field(default_factory=list)

@dataclass
class SourceCache:
    source: str
    loaded_at: float
    workbook_bytes: bytes
    frames: Dict[str, pd.DataFrame] = field(default_factory=dict)
    indices: Dict[str, ValueIndex] = field(default_factory=dict)
    lock: threading.Lock = field(default_factory=threading.Lock)

_cache: Dict[str, SourceCache] = {}
_cache_lock = threading.Lock()

app = FastAPI(title="AMPM Analytical Chat Engine", version=APP_VERSION)
origins_raw = os.getenv("ALLOWED_ORIGINS", "*").strip()
allow_origins = [o.strip() for o in origins_raw.split(",") if o.strip()] or ["*"]
app.add_middleware(CORSMiddleware, allow_origins=allow_origins, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


def norm(value: Any) -> str:
    raw = "" if value is None else str(value)
    raw = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", raw.lower()).strip()


def txt(value: Any) -> str:
    return "" if value is None else str(value).strip()


def fmt_money(value: float) -> str:
    return f"${float(value):,.2f}"


def fmt_num(value: float) -> str:
    return f"{float(value):,.0f}"


def fmt_pct(value: Optional[float]) -> str:
    if value is None:
        return "N/D"
    try:
        val = float(value)
    except Exception:
        return "N/D"
    if not math.isfinite(val):
        return "N/D"
    return f"{val * 100:.1f}%"


def pct_delta(current: float, base: float) -> Optional[float]:
    if base in (None, 0) or pd.isna(base):
        return None
    return (float(current) / float(base)) - 1.0


def month_label(month_key: str) -> str:
    try:
        year, month = month_key.split("-")
        return f"{MONTHS_ABBR[int(month)]} {year}"
    except Exception:
        return month_key


def same_month_ly(month_key: str) -> Optional[str]:
    try:
        year, month = month_key.split("-")
        return f"{int(year)-1:04d}-{month}"
    except Exception:
        return None


def previous_month(month_key: str, months: List[str]) -> Optional[str]:
    if month_key not in months:
        return None
    idx = months.index(month_key)
    return months[idx - 1] if idx > 0 else None


def dedupe_keep_order(values: Iterable[str]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            out.append(value)
    return out


def read_sheet_df(ws: openpyxl.worksheet.worksheet.Worksheet, columns: List[str]) -> pd.DataFrame:
    rows = ws.iter_rows(values_only=True)
    header = ["" if c is None else str(c).strip() for c in next(rows)]
    positions = [header.index(col) for col in columns]
    data: List[List[Any]] = []
    for row in rows:
        data.append([row[pos] if pos < len(row) else None for pos in positions])
    return pd.DataFrame(data, columns=columns)


def prepare_frame(key: str, df: pd.DataFrame) -> pd.DataFrame:
    cfg = SHEET_CONFIG[key]
    df = df.copy()
    for col in cfg.get("dates", []):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    for col in cfg.get("text", []):
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()
    for col in cfg.get("numeric", []):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    if "Is_Delivery" in df.columns:
        df["Is_Delivery"] = df["Is_Delivery"].apply(lambda v: True if v in (True, 1, "1", "true", "True", "TRUE") else False)
    return df


def get_source_cache(source: str) -> SourceCache:
    now = time.time()
    with _cache_lock:
        cached = _cache.get(source)
        if cached and now - cached.loaded_at <= CACHE_TTL_SECONDS:
            return cached
    if source.startswith("http://") or source.startswith("https://"):
        response = requests.get(source, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        workbook_bytes = response.content
    else:
        if not os.path.exists(source):
            raise FileNotFoundError(source)
        with open(source, "rb") as fh:
            workbook_bytes = fh.read()
    cache = SourceCache(source=source, loaded_at=now, workbook_bytes=workbook_bytes)
    with _cache_lock:
        _cache[source] = cache
    return cache


def get_frame(cache: SourceCache, key: str) -> pd.DataFrame:
    if key in cache.frames:
        return cache.frames[key]
    with cache.lock:
        if key in cache.frames:
            return cache.frames[key]
        cfg = SHEET_CONFIG[key]
        wb = openpyxl.load_workbook(io.BytesIO(cache.workbook_bytes), read_only=True, data_only=True)
        sheet_name = cfg["sheet"]
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"No encontré la hoja requerida: {sheet_name}")
        df = prepare_frame(key, read_sheet_df(wb[sheet_name], cfg["columns"]))
        cache.frames[key] = df
        return df


def get_index(cache: SourceCache, key: str) -> ValueIndex:
    if key in cache.indices:
        return cache.indices[key]
    with cache.lock:
        if key in cache.indices:
            return cache.indices[key]
        df = get_frame(cache, key)
        cfg = SHEET_CONFIG[key]
        values_by_col: Dict[str, List[Tuple[str, List[str]]]] = {}
        for col in cfg.get("text", []):
            if col not in df.columns:
                continue
            buckets: Dict[str, List[str]] = {}
            for original in df[col].dropna().astype(str).unique().tolist():
                cleaned = txt(original)
                if not cleaned:
                    continue
                normalized = norm(cleaned)
                if len(normalized) < 2:
                    continue
                buckets.setdefault(normalized, []).append(cleaned)
            values_by_col[col] = sorted(buckets.items(), key=lambda item: (-len(item[0]), item[0]))
        period_col = "CalMonth" if "CalMonth" in df.columns else "Month" if "Month" in df.columns else None
        months = sorted([m for m in df[period_col].dropna().astype(str).unique().tolist() if period_col and re.match(r"^\d{4}-\d{2}$", str(m))]) if period_col else []
        stores = sorted(df["Store"].dropna().astype(str).unique().tolist()) if "Store" in df.columns else []
        idx = ValueIndex(values_by_col=values_by_col, months=months, stores=stores)
        cache.indices[key] = idx
        return idx


def extract_months_in_order(qn: str, months: List[str]) -> List[str]:
    matches: List[Tuple[int, str]] = []
    for name, month_num in MONTHS_ES.items():
        for match in re.finditer(rf"\b{name}\b(?:\s+de)?\s*(20\d{{2}})?", qn):
            year = match.group(1)
            if year:
                candidate = f"{year}-{month_num:02d}"
                if candidate in months:
                    matches.append((match.start(), candidate))
            else:
                candidates = [m for m in months if m.endswith(f"-{month_num:02d}")]
                if candidates:
                    matches.append((match.start(), candidates[-1]))
    for match in re.finditer(r"\b(20\d{2})[-/](0[1-9]|1[0-2])\b", qn):
        candidate = f"{match.group(1)}-{match.group(2)}"
        if candidate in months:
            matches.append((match.start(), candidate))
    ordered = dedupe_keep_order([candidate for _, candidate in sorted(matches, key=lambda item: item[0])])
    return ordered


def extract_target_month(qn: str, months: List[str], filters: Dict[str, Any]) -> Optional[str]:
    ordered = extract_months_in_order(qn, months)
    if ordered:
        return ordered[0]
    current = txt(filters.get("current_period"))
    if current and current in months:
        return current
    return months[-1] if months else None


def extract_store_filters(qn: str, filters: Dict[str, Any], valid_stores: List[str]) -> List[str]:
    hits: List[str] = []
    for match in re.findall(r"\b(?:ampm\s*0?([1-9])|a0?([1-9]))\b", qn):
        digit = next((piece for piece in match if piece), None)
        if digit:
            hits.append(f"AMPM{int(digit):02d}")
    for raw in filters.get("stores") or []:
        value = txt(raw).upper()
        if value:
            hits.append(value)
    valid = set(valid_stores)
    return sorted({store for store in hits if store in valid})


def has_any(qn: str, *terms: str) -> bool:
    return any(term in qn for term in terms)


def wants_chart(qn: str) -> bool:
    return has_any(qn, "graf", "chart", "barra", "barras", "pastel", "pie", "donut", "doughnut", "visual", "muestrame", "muéstrame")


def chart_type(qn: str) -> str:
    if has_any(qn, "pastel", "pie", "donut", "doughnut"):
        return "pie"
    if has_any(qn, "linea", "line", "tendencia", "evolucion", "evolución"):
        return "line"
    return "bar"


def top_n(qn: str) -> int:
    match = re.search(r"\btop\s*(\d{1,2})\b", qn)
    if match:
        return max(1, min(20, int(match.group(1))))
    return 10 if "top" in qn else 1


def compare_mode(qn: str, months_mentioned: List[str]) -> str:
    if len(months_mentioned) >= 2:
        return "pair"
    if has_any(qn, "ly", "año pasado", "ano pasado", "vs ly", "contra ly", "delta vs ly"):
        return "ly"
    if has_any(qn, "mes pasado", "lm", "vs mes pasado", "contra febrero", "contra marzo", "contra abril"):
        return "previous"
    if has_any(qn, "decrec", "cayo", "cayó", "baja", "sube", "crece", "cae"):
        return "previous"
    return "none"


def domain_scores(qn: str, active_tab: str) -> Dict[str, int]:
    scores = {key: 0 for key in DOMAIN_LABELS}
    def add(domain: str, points: int, *terms: str) -> None:
        for term in terms:
            if term in qn:
                scores[domain] += points
    add("delivery", 9, "delivery", "pedido ya", "pedidos ya", "pedidoya", "uber", "canal")
    add("sbf", 10, "sbf", "puntoxpress", "aki pago", "western union", "recarga", "rapibac", "smart ticket", "servicio", "servicios")
    add("inventory", 10, "inventario", "stock", "doi", "dias inv", "días inv", "planimetria", "targetqty", "idealqty")
    add("merma", 10, "merma", "desperdicio", "vencimiento")
    add("innovation", 10, "innovacion", "innovación", "food service", "combo", "hot dog", "pizza", "hamburg", "desayuno")
    add("cxc", 10, "cxc", "faltante", "sobrante", "caja", "batch", "cashier")
    add("prd0", 10, "prd en cero", "prd0", "quiebre", "agotado", "lost sales", "venta perdida")
    add("hallazgos", 9, "hallazgo", "hallazgos", "lectura ejecutiva", "riesgo", "oportunidad", "acciones")
    add("sales", 3, "ventas", "venta", "sales", "categoria", "categoría", "marca", "marcas", "producto", "productos", "descripcion", "descripción")
    if active_tab in scores:
        scores[active_tab] += 2
    return scores


def pick_domain(qn: str, active_tab: str) -> Tuple[str, List[str]]:
    ranked = sorted(domain_scores(qn, active_tab).items(), key=lambda item: item[1], reverse=True)
    explicit = [name for name, score in ranked if score >= 9]
    if len(explicit) >= 2:
        return explicit[0], explicit[:2]
    if ranked and ranked[0][1] > 0:
        return ranked[0][0], []
    return (active_tab if active_tab in DOMAIN_LABELS else "sales"), []


def match_entities(qn: str, cache: SourceCache, key: str, columns: List[str]) -> Dict[str, List[str]]:
    idx = get_index(cache, key)
    hits: Dict[str, List[str]] = {}
    for col in columns:
        values = idx.values_by_col.get(col, [])
        matched: List[str] = []
        for normalized, originals in values:
            if len(normalized) < 3:
                continue
            if re.search(rf"(?<!\w){re.escape(normalized)}(?!\w)", qn):
                matched.extend(originals)
        if matched:
            hits[col] = dedupe_keep_order(matched)
    return hits


def explicit_dimension_mentions(qn: str) -> Dict[str, bool]:
    return {
        "Category": has_any(qn, "categoria", "categoría", "category"),
        "Brand": has_any(qn, "marca", "marcas", "brand"),
        "Description": has_any(qn, "producto", "productos", "sku", "descripcion", "descripción", "item"),
        "Department": has_any(qn, "departamento", "department"),
        "Supplier": has_any(qn, "proveedor", "supplier"),
        "Provider": has_any(qn, "servicio", "servicios", "proveedor", "provider"),
        "ComboLabel": has_any(qn, "combo", "combos"),
    }


def infer_group(qn: str, domain: str, entity_hits: Dict[str, List[str]]) -> Optional[str]:
    mentions = explicit_dimension_mentions(qn)
    if domain == "sbf":
        if mentions["Provider"]:
            return "Provider"
        if has_any(qn, "tienda", "tiendas", "store"):
            return "Store"
        if has_any(qn, "hora", "horas"):
            return "Hour"
        if has_any(qn, "horario", "turno", "daypart"):
            return "Daypart"
        if has_any(qn, "dia de semana", "día de semana", "dow"):
            return "DOW"
        return "Store" if wants_chart(qn) else None
    if mentions["Description"]:
        return "Description"
    if mentions["Brand"]:
        return "Brand"
    if mentions["Category"]:
        if entity_hits.get("Category") and has_any(qn, "producto", "productos", "mas vendido", "más vendido", "top", "lider", "líder"):
            return "Description"
        return "Category"
    if has_any(qn, "tienda", "tiendas", "store"):
        return "Store"
    if has_any(qn, "hora", "horas"):
        return "Hour"
    if has_any(qn, "horario", "turno", "daypart"):
        return "Daypart"
    if has_any(qn, "dia de semana", "día de semana", "dow"):
        return "DOW_Name"
    if has_any(qn, "canal", "channel"):
        return "Delivery_Channel"
    if domain == "innovation" and has_any(qn, "combo", "combos"):
        return "ComboLabel"
    return None


def metric_for(domain: str, qn: str) -> str:
    if domain in {"sales", "delivery", "innovation"}:
        if has_any(qn, "cantidad", "qty", "unidades"):
            return "Qty Sold"
        if has_any(qn, "margen", "margin"):
            return "Total Gross Margin"
        return "Sales"
    if domain == "sbf":
        return "TX"
    if domain == "inventory":
        if has_any(qn, "dias inv", "días inv", "doi"):
            return "Días Inv"
        if has_any(qn, "cantidad", "qty", "unidades"):
            return "Quantity"
        if has_any(qn, "sales s4", "venta s4"):
            return "Sales S4"
        return "Amount"
    if domain == "merma":
        if has_any(qn, "packs", "paquetes"):
            return "Quantity (packs)"
        if has_any(qn, "sales", "venta") and "merma" not in qn:
            return "Total Sales"
        return "Merma"
    if domain == "cxc":
        return "Amount"
    if domain == "prd0":
        if has_any(qn, "lost units", "unidades perdidas"):
            return "Lost Units 3d"
        if has_any(qn, "rop"):
            return "ROP"
        return "Lost Sales 3d"
    return "Sales"


def apply_common_filters(df: pd.DataFrame, stores: List[str], month_key: Optional[str], filters: Dict[str, Any]) -> pd.DataFrame:
    out = df.copy()
    if stores and "Store" in out.columns:
        out = out[out["Store"].isin(stores)].copy()
    period_col = "CalMonth" if "CalMonth" in out.columns else "Month" if "Month" in out.columns else None
    if month_key and period_col:
        out = out[out[period_col].astype(str) == month_key].copy()
    day_filter = filters.get("day_of_week") or []
    if day_filter and "DOW" in out.columns:
        values: List[int] = []
        for raw in day_filter:
            try:
                values.append(int(raw))
            except Exception:
                continue
        if values:
            out = out[out["DOW"].isin(values)].copy()
    return out


def apply_sales_domain_filter(df: pd.DataFrame, domain: str) -> pd.DataFrame:
    if domain == "delivery" and "Is_Delivery" in df.columns:
        return df[df["Is_Delivery"] == True].copy()  # noqa: E712
    return df


def apply_named_filters(df: pd.DataFrame, qn: str, domain: str, entity_hits: Dict[str, List[str]], group_col: Optional[str]) -> Tuple[pd.DataFrame, List[str]]:
    out = df.copy()
    applied: List[str] = []
    mentions = explicit_dimension_mentions(qn)
    for col, values in entity_hits.items():
        if col not in out.columns or not values:
            continue
        should_filter = False
        if col != group_col:
            should_filter = True
        elif group_col == "Category" and has_any(qn, "de la categoria", "de categoria", "en la categoria", "en categoria"):
            should_filter = True
        elif group_col == "Brand" and has_any(qn, "de la marca", "en la marca"):
            should_filter = True
        elif group_col == "Description" and has_any(qn, "de la descripcion", "de la descripción"):
            should_filter = True
        elif col == "Category" and mentions["Description"]:
            should_filter = True
        elif col == "Brand" and mentions["Description"]:
            should_filter = True
        elif col == "Provider" and group_col != "Provider":
            should_filter = True
        elif col == "ComboLabel" and group_col != "ComboLabel":
            should_filter = True
        if should_filter:
            out = out[out[col].isin(values)].copy()
            if values:
                applied.append(f"{col}: {', '.join(values[:3])}")
    if domain in {"sales", "delivery", "innovation", "inventory", "merma", "prd0"}:
        if has_any(qn, "cigarro", "cigarros", "cigarrillo", "cigarrillos") and "Category" in out.columns:
            mask = out["Category"].astype(str).map(norm).str.contains("cigarr")
            if mask.any():
                out = out[mask].copy()
                applied.append("Familia: Cigarrillos")
        elif has_any(qn, "vape", "vapes") and "Category" in out.columns:
            mask = out["Category"].astype(str).map(norm).str.contains("vape")
            if mask.any():
                out = out[mask].copy()
                applied.append("Familia: Vapes")
        elif has_any(qn, "licor", "licores") and "Category" in out.columns:
            mask = out["Category"].astype(str).map(norm).str.contains("licor|vino|champ")
            if mask.any():
                out = out[mask].copy()
                applied.append("Familia: Licores")
    if domain == "delivery" and "Delivery_Channel" in out.columns:
        if has_any(qn, "pedido ya", "pedidos ya", "pedidoya", "peya"):
            mask = out["Delivery_Channel"].astype(str).map(norm).str.contains("pedido")
            if mask.any():
                out = out[mask].copy()
                applied.append("Canal: Pedidos Ya")
        elif "uber" in qn:
            mask = out["Delivery_Channel"].astype(str).map(norm).str.contains("uber")
            if mask.any():
                out = out[mask].copy()
                applied.append("Canal: Uber")
    return out, applied


def build_chart(labels: List[str], datasets: List[Dict[str, Any]], title: str, subtitle: str, qn: str) -> Dict[str, Any]:
    return {
        "title": title,
        "subtitle": subtitle,
        "type": chart_type(qn),
        "labels": labels,
        "datasets": datasets,
        "note": "Fuente: Cubo_Semanal_Compactado.xlsx",
    }


def summarise_top(df: pd.DataFrame, label_col: str, value_col: str, limit: int = 5) -> str:
    items: List[str] = []
    for _, row in df.head(limit).iterrows():
        label = txt(row[label_col]) or "Sin etiqueta"
        value = float(row[value_col] or 0)
        rendered = fmt_num(value) if value_col in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(value)
        items.append(f"{label}: {rendered}")
    return " | ".join(items)


def default_actions(domain: str, negative: bool = False) -> List[str]:
    if domain == "sales":
        return [
            "Aterriza el seguimiento en la línea con mayor peso o mayor caída absoluta.",
            "Si el delta cae, revisa disponibilidad, precio y mix antes de abrir campaña.",
            "Replica la ejecución de la tienda o marca que sí sostiene el resultado.",
        ]
    if domain == "delivery":
        return [
            "Concentra visibilidad y surtido en el canal y franja que sí convierten.",
            "Si LY cae, revisa disponibilidad y tiempos por tienda antes de empujar pauta.",
            "Usa la tienda top como benchmark operativo del canal.",
        ]
    if domain == "sbf":
        return [
            "Protege cobertura en el servicio y tienda con mayor peso transaccional.",
            "Si el delta cae, revisa visibilidad del servicio y disciplina de caja.",
            "Replica horarios ganadores donde la TX ya demuestra tracción.",
        ]
    if domain == "inventory":
        return [
            "Prioriza exceso y días de inventario altos en las líneas de menor rotación.",
            "Evita recompras donde la cobertura ya está sobrada.",
            "Protege disponibilidad en SKU de alta venta S4 antes de mover capital al resto.",
        ]
    if domain == "merma":
        return [
            "Ataca primero las descripciones con mayor impacto económico absoluto.",
            "Cruza merma con venta para separar ruido de problema material.",
            "Corrige causa raíz en tienda antes de ampliar pedido del mismo SKU.",
        ]
    if domain == "innovation":
        return [
            "Escala el combo o bloque que ya muestra tracción real.",
            "Si un bloque no rota, corrige surtido o comunicación antes de ampliar portafolio.",
            "Usa el mix top como base del siguiente test comercial.",
        ]
    if domain == "cxc":
        return [
            "Ataca primero el batch o comentario con mayor monto absoluto.",
            "Separa eventos compensados de faltantes reales antes de decidir.",
            "Refuerza disciplina de caja en la tienda con mayor recurrencia.",
        ]
    if domain == "prd0":
        return [
            "Prioriza los SKU con mayor venta perdida estimada en 3 días.",
            "Corrige ROP y compra en líneas de alta rotación.",
            "Usa reposición diaria en los quiebres realmente materiales.",
        ]
    return [
        "Convierte el hallazgo principal en una acción operativa y una comercial.",
        "Asigna dueño, tienda y fecha de seguimiento para evitar que quede declarativo.",
        "Valida el mismo corte en la siguiente revisión para confirmar impacto.",
    ]


def lookup_group_label(group_col: str) -> str:
    labels = {
        "Store": "tienda",
        "Category": "categoría",
        "Brand": "marca",
        "Description": "producto",
        "Department": "departamento",
        "Supplier": "proveedor",
        "Provider": "servicio",
        "Daypart": "horario",
        "DOW_Name": "día de semana",
        "DOW": "día de semana",
        "Hour": "hora",
        "Delivery_Channel": "canal",
        "ComboLabel": "combo",
    }
    return labels.get(group_col, group_col)


def normalize_group_labels(df: pd.DataFrame, group_col: str) -> pd.DataFrame:
    out = df.copy()
    if group_col == "DOW" and "DOW" in out.columns:
        out["DOW"] = out["DOW"].map(lambda v: DOW_NUM_TO_ES.get(int(v), f"Día {int(v)}") if float(v).is_integer() else str(v))
    return out


def aggregate_group(df: pd.DataFrame, group_col: str, metric: str) -> pd.DataFrame:
    if group_col not in df.columns:
        raise RuntimeError(f"La dimensión solicitada no existe en la fuente: {group_col}")
    grouped = df.groupby(group_col, dropna=False)[metric].sum().reset_index()
    grouped[group_col] = grouped[group_col].fillna("").astype(str).str.strip()
    grouped = grouped[grouped[group_col] != ""]
    return normalize_group_labels(grouped, group_col)


def prepare_compare_group(current_df: pd.DataFrame, base_df: pd.DataFrame, group_col: str, metric: str) -> pd.DataFrame:
    cur = aggregate_group(current_df, group_col, metric).rename(columns={metric: "current"})
    base = aggregate_group(base_df, group_col, metric).rename(columns={metric: "base"})
    merged = cur.merge(base, on=group_col, how="outer").fillna(0)
    merged["delta_abs"] = merged["current"] - merged["base"]
    merged["delta_pct"] = merged.apply(lambda row: pct_delta(row["current"], row["base"]), axis=1)
    return merged


def driver_table(current_df: pd.DataFrame, base_df: pd.DataFrame, metric: str, preferred: List[str]) -> Optional[pd.DataFrame]:
    for group_col in preferred:
        if group_col in current_df.columns and current_df[group_col].astype(str).str.strip().nunique() > 1:
            table = prepare_compare_group(current_df, base_df, group_col, metric)
            if not table.empty:
                return table.rename(columns={group_col: "label"}).assign(group_col=group_col)
    return None


def build_notes(cache: SourceCache, month_key: Optional[str], stores: List[str], applied_filters: List[str]) -> List[str]:
    notes: List[str] = []
    if month_key:
        notes.append(f"Periodo: {month_label(month_key)}")
    if stores:
        notes.append(f"Tiendas: {', '.join(stores)}")
    if applied_filters:
        notes.append(f"Filtros: {' · '.join(applied_filters[:4])}")
    notes.append(f"Fuente de verdad del chat: {cache.source}")
    return notes


def resolve_rank_or_total(cache: SourceCache, domain: str, df: pd.DataFrame, qn: str, month_key: Optional[str], stores: List[str], metric: str, group_col: Optional[str], applied_filters: List[str]) -> Dict[str, Any]:
    title_base = DOMAIN_LABELS[domain]
    if not group_col:
        total = float(df[metric].sum())
        rendered = fmt_num(total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(total)
        return {
            "title": f"{title_base} total",
            "summary": f"Corte {month_label(month_key) if month_key else 'actual'}. Total visible: {rendered}.",
            "actions": default_actions(domain),
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": None,
        }
    grouped = aggregate_group(df, group_col, metric).sort_values(metric, ascending=False)
    n = top_n(qn)
    top = grouped.head(n if n > 1 else 10)
    total = float(df[metric].sum())
    metric_render = fmt_num(total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(total)
    group_label = lookup_group_label(group_col)
    if n == 1 and has_any(qn, "cual es", "cuál es", "mas vendido", "más vendido", "lider", "líder", "top"):
        winner = top.iloc[0]
        winner_label = txt(winner[group_col])
        winner_value = float(winner[metric])
        winner_render = fmt_num(winner_value) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(winner_value)
        chart = build_chart([txt(v) for v in top[group_col].tolist()], [{"label": metric, "data": [float(v) for v in top[metric].tolist()]}], f"{title_base} por {group_label}", f"Corte {month_label(month_key) if month_key else 'actual'}", qn)
        return {
            "title": f"{title_base} · {group_label} líder",
            "summary": f"En {month_label(month_key) if month_key else 'el corte visible'}, el {group_label} líder es {winner_label} con {winner_render}. Visible total: {metric_render}. Siguiente bloque visible: {summarise_top(top.iloc[1:], group_col, metric, limit=4) if len(top) > 1 else 'sin siguientes materiales' }.",
            "actions": default_actions(domain),
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": chart if wants_chart(qn) or "top" in qn or "mas vendido" in qn or "más vendido" in qn else None,
        }
    chart = build_chart([txt(v) for v in top[group_col].tolist()], [{"label": metric, "data": [float(v) for v in top[metric].tolist()]}], f"{title_base} por {group_label}", f"Corte {month_label(month_key) if month_key else 'actual'}", qn)
    return {
        "title": f"{title_base} por {group_label}",
        "summary": f"Corte {month_label(month_key) if month_key else 'actual'}. Total visible: {metric_render}. Top visible: {summarise_top(top, group_col, metric)}.",
        "actions": default_actions(domain),
        "notes": build_notes(cache, month_key, stores, applied_filters),
        "chart": chart if wants_chart(qn) or n > 1 or "top" in qn else None,
    }


def resolve_compare(cache: SourceCache, domain: str, current_df: pd.DataFrame, base_df: pd.DataFrame, qn: str, month_key: str, base_month: str, stores: List[str], metric: str, group_col: Optional[str], applied_filters: List[str], explicit_explain: bool) -> Dict[str, Any]:
    title_base = DOMAIN_LABELS[domain]
    current_total = float(current_df[metric].sum())
    base_total = float(base_df[metric].sum())
    delta_abs = current_total - base_total
    delta_pct = pct_delta(current_total, base_total)
    baseline_label = month_label(base_month)
    current_label = month_label(month_key)
    negative_signal = delta_abs < 0

    if not group_col:
        default_group = {
            "sales": "Category",
            "delivery": "Store",
            "sbf": "Store",
            "inventory": "Store",
            "merma": "Description",
            "innovation": "ComboLabel",
            "cxc": "Store",
            "prd0": "Description",
        }
        group_col = default_group.get(domain)

    table = prepare_compare_group(current_df, base_df, group_col, metric) if group_col else pd.DataFrame()
    table = table.sort_values("current", ascending=False) if not table.empty else table
    chart = None
    if not table.empty:
        top = table.head(10)
        chart = build_chart(
            [txt(v) for v in top[group_col].tolist()],
            [
                {"label": current_label, "data": [float(v) for v in top["current"].tolist()]},
                {"label": baseline_label, "data": [float(v) for v in top["base"].tolist()]},
            ],
            f"{title_base} por {lookup_group_label(group_col)}",
            f"{current_label} vs {baseline_label}",
            qn,
        )

    if explicit_explain:
        drivers = driver_table(current_df, base_df, metric, ["Store", "Brand", "Description", "Category", "Daypart", "Provider"])
        driver_text = ""
        if drivers is not None and not drivers.empty:
            falling = drivers.sort_values("delta_abs").head(3)
            if delta_abs < 0:
                driver_text = " Principales arrastres visibles: " + " | ".join(
                    f"{txt(row['label'])}: {fmt_money(float(row['delta_abs'])) if metric not in {'TX', 'Quantity', 'Qty Sold', 'ROP', 'Lost Units 3d', 'Días Inv', 'Quantity (packs)'} else fmt_num(float(row['delta_abs']))}"
                    for _, row in falling.iterrows()
                ) + "."
            else:
                winners = drivers.sort_values("delta_abs", ascending=False).head(3)
                driver_text = " El crecimiento visible se explica sobre todo por: " + " | ".join(
                    f"{txt(row['label'])}: {fmt_money(float(row['delta_abs'])) if metric not in {'TX', 'Quantity', 'Qty Sold', 'ROP', 'Lost Units 3d', 'Días Inv', 'Quantity (packs)'} else fmt_num(float(row['delta_abs']))}"
                    for _, row in winners.iterrows()
                ) + "."
        rendered_current = fmt_num(current_total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(current_total)
        rendered_base = fmt_num(base_total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(base_total)
        direction_line = (
            f"En {current_label}, el visible cae {fmt_money(abs(delta_abs)) if metric not in {'TX', 'Quantity', 'Qty Sold', 'ROP', 'Lost Units 3d', 'Días Inv', 'Quantity (packs)'} else fmt_num(abs(delta_abs))} vs {baseline_label} ({fmt_pct(delta_pct)})."
            if delta_abs < 0 else
            f"En {current_label}, no veo decrecimiento frente a {baseline_label}; el visible sube {fmt_money(delta_abs) if metric not in {'TX', 'Quantity', 'Qty Sold', 'ROP', 'Lost Units 3d', 'Días Inv', 'Quantity (packs)'} else fmt_num(delta_abs)} ({fmt_pct(delta_pct)})."
        )
        summary = f"{direction_line} Actual: {rendered_current} vs base {rendered_base}.{driver_text}"
        return {
            "title": f"{title_base} · lectura de variación",
            "summary": summary,
            "actions": default_actions(domain, negative=negative_signal),
            "notes": build_notes(cache, month_key, stores, applied_filters + [f"Base comparativa: {baseline_label}"]),
            "chart": chart if wants_chart(qn) or not table.empty else None,
        }

    top_text = summarise_top(table.head(5).rename(columns={"current": metric}), group_col, metric) if not table.empty else "sin aperturas materiales"
    rendered_current = fmt_num(current_total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(current_total)
    rendered_base = fmt_num(base_total) if metric in {"TX", "Quantity", "Qty Sold", "ROP", "Lost Units 3d", "Días Inv", "Quantity (packs)"} else fmt_money(base_total)
    return {
        "title": f"{title_base} · vs {baseline_label}",
        "summary": f"Periodo {current_label}. Actual: {rendered_current} vs base {rendered_base}. Delta: {fmt_pct(delta_pct)}. Top visible: {top_text}.",
        "actions": default_actions(domain, negative=negative_signal),
        "notes": build_notes(cache, month_key, stores, applied_filters + [f"Base comparativa: {baseline_label}"]),
        "chart": chart,
    }


def resolve_sales_like(cache: SourceCache, domain: str, payload: QueryRequest) -> Dict[str, Any]:
    qn = norm(payload.question)
    sales_df = get_frame(cache, "sales") if domain in {"sales", "delivery"} else get_frame(cache, "innovation")
    idx = get_index(cache, "sales" if domain in {"sales", "delivery"} else "innovation")
    month_keys = idx.months
    month_mentions = extract_months_in_order(qn, month_keys)
    month_key = extract_target_month(qn, month_keys, payload.filters)
    stores = extract_store_filters(qn, payload.filters, idx.stores)
    df = apply_sales_domain_filter(sales_df, domain)
    df = apply_common_filters(df, stores, month_key, payload.filters)
    entity_hits = match_entities(qn, cache, "sales" if domain in {"sales", "delivery"} else "innovation", ["Department", "Category", "Supplier", "Brand", "Description", "ComboLabel", "Daypart", "Delivery_Channel"])
    group_col = infer_group(qn, domain, entity_hits)
    df, applied_filters = apply_named_filters(df, qn, domain, entity_hits, group_col)
    if has_any(qn, "hora", "horas") and "Hour" not in df.columns:
        return {
            "title": f"{DOMAIN_LABELS[domain]} · hora no disponible",
            "summary": "La fuente actual no trae hora real para este dominio. Sí trae día de semana y horario operativo, así que la consulta correcta es por día de semana o por horario.",
            "actions": [
                "Pídelo por día de semana para tener una lectura utilizable ya.",
                "Pídelo por horario operativo para capturar la señal comercial disponible.",
                "Si quieres hora real, agrega Hour a la fuente de este dominio.",
            ],
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": None,
        }
    if df.empty:
        return {
            "title": f"{DOMAIN_LABELS[domain]} · sin data visible",
            "summary": "No encontré filas visibles con la combinación actual de periodo, tienda y filtros de la pregunta.",
            "actions": [
                "Revisa si la tienda o el mes pedido existe en la fuente publicada.",
                "Prueba la misma pregunta sin filtro de tienda para validar cobertura.",
                "Si esperabas un grano más fino, revisa si esa dimensión existe en el workbook.",
            ],
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": None,
        }
    metric = metric_for(domain, qn)
    mode = compare_mode(qn, month_mentions)
    explicit_explain = has_any(qn, "por que", "por qué", "explica", "motivo", "razon", "razón")

    if mode == "pair" and len(month_mentions) >= 2:
        current_month, base_month = month_mentions[0], month_mentions[1]
    elif mode == "ly":
        current_month, base_month = month_key, same_month_ly(month_key) if month_key else None
    elif mode == "previous":
        current_month, base_month = month_key, previous_month(month_key, month_keys) if month_key else None
    else:
        current_month, base_month = month_key, None

    if base_month:
        current_df = apply_sales_domain_filter(sales_df, domain)
        current_df = apply_common_filters(current_df, stores, current_month, payload.filters)
        current_df, _ = apply_named_filters(current_df, qn, domain, entity_hits, group_col)
        base_df = apply_sales_domain_filter(sales_df, domain)
        base_df = apply_common_filters(base_df, stores, base_month, payload.filters)
        base_df, _ = apply_named_filters(base_df, qn, domain, entity_hits, group_col)
        if not current_df.empty and not base_df.empty:
            return resolve_compare(cache, domain, current_df, base_df, qn, current_month, base_month, stores, metric, group_col, applied_filters, explicit_explain or mode != "none")

    if group_col is None and (wants_chart(qn) or "top" in qn or has_any(qn, "mas vendido", "más vendido", "lider", "líder")):
        group_col = "Description" if entity_hits.get("Category") else "Category"
    return resolve_rank_or_total(cache, domain, df, qn, month_key, stores, metric, group_col, applied_filters)


def resolve_sbf(cache: SourceCache, payload: QueryRequest) -> Dict[str, Any]:
    qn = norm(payload.question)
    df_all = get_frame(cache, "sbf")
    idx = get_index(cache, "sbf")
    month_mentions = extract_months_in_order(qn, idx.months)
    month_key = extract_target_month(qn, idx.months, payload.filters)
    stores = extract_store_filters(qn, payload.filters, idx.stores)
    entity_hits = match_entities(qn, cache, "sbf", ["Provider", "Daypart"])
    group_col = infer_group(qn, "sbf", entity_hits)
    metric = "TX"

    current_df = apply_common_filters(df_all, stores, month_key, payload.filters)
    current_df, applied_filters = apply_named_filters(current_df, qn, "sbf", entity_hits, group_col)
    if group_col == "DOW" and "DOW" in current_df.columns:
        current_df = current_df.copy()
        current_df["DOW"] = current_df["DOW"].map(lambda v: DOW_NUM_TO_ES.get(int(v), str(v)) if not pd.isna(v) else "")
    if current_df.empty:
        return {
            "title": "SBF · sin data visible",
            "summary": "No encontré TX visibles con la combinación actual de periodo, tienda y filtros de la pregunta.",
            "actions": default_actions("sbf"),
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": None,
        }

    mode = compare_mode(qn, month_mentions)
    explicit_explain = has_any(qn, "por que", "por qué", "explica", "motivo", "razon", "razón")
    if mode == "pair" and len(month_mentions) >= 2:
        current_month, base_month = month_mentions[0], month_mentions[1]
    elif mode == "ly":
        current_month, base_month = month_key, same_month_ly(month_key) if month_key else None
    elif mode == "previous":
        current_month, base_month = month_key, previous_month(month_key, idx.months) if month_key else None
    else:
        current_month, base_month = month_key, None
    if base_month:
        current_df = apply_common_filters(df_all, stores, current_month, payload.filters)
        current_df, _ = apply_named_filters(current_df, qn, "sbf", entity_hits, group_col)
        base_df = apply_common_filters(df_all, stores, base_month, payload.filters)
        base_df, _ = apply_named_filters(base_df, qn, "sbf", entity_hits, group_col)
        if group_col == "DOW":
            current_df = current_df.copy(); current_df["DOW"] = current_df["DOW"].map(lambda v: DOW_NUM_TO_ES.get(int(v), str(v)) if not pd.isna(v) else "")
            base_df = base_df.copy(); base_df["DOW"] = base_df["DOW"].map(lambda v: DOW_NUM_TO_ES.get(int(v), str(v)) if not pd.isna(v) else "")
        if not current_df.empty and not base_df.empty:
            return resolve_compare(cache, "sbf", current_df, base_df, qn, current_month, base_month, stores, metric, group_col, applied_filters, explicit_explain or mode != "none")

    if group_col is None and (wants_chart(qn) or "top" in qn):
        group_col = "Store"
    return resolve_rank_or_total(cache, "sbf", current_df, qn, month_key, stores, metric, group_col, applied_filters)


def resolve_inventory_merma_cxc_prd0(cache: SourceCache, domain: str, payload: QueryRequest) -> Dict[str, Any]:
    qn = norm(payload.question)
    key = domain if domain != "cxc" else "cxc"
    df_all = get_frame(cache, key)
    idx = get_index(cache, key)
    month_mentions = extract_months_in_order(qn, idx.months)
    month_key = extract_target_month(qn, idx.months, payload.filters)
    stores = extract_store_filters(qn, payload.filters, idx.stores)
    cols = [col for col in ["Department", "Category", "Supplier", "Brand", "Description", "Comment", "Cashier"] if col in df_all.columns]
    entity_hits = match_entities(qn, cache, key, cols)
    group_col = infer_group(qn, domain, entity_hits)
    metric = metric_for(domain, qn)

    df = apply_common_filters(df_all, stores, month_key, payload.filters)
    df, applied_filters = apply_named_filters(df, qn, domain, entity_hits, group_col)
    if df.empty:
        return {
            "title": f"{DOMAIN_LABELS[domain]} · sin data visible",
            "summary": "No encontré filas visibles con la combinación actual de periodo, tienda y filtros de la pregunta.",
            "actions": default_actions(domain),
            "notes": build_notes(cache, month_key, stores, applied_filters),
            "chart": None,
        }
    if group_col is None and (wants_chart(qn) or "top" in qn or has_any(qn, "mas vendido", "más vendido")):
        defaults = {"inventory": "Store", "merma": "Description", "cxc": "Store", "prd0": "Description"}
        group_col = defaults.get(domain)
    return resolve_rank_or_total(cache, domain, df, qn, month_key, stores, metric, group_col, applied_filters)


def resolve_hallazgos(cache: SourceCache, payload: QueryRequest) -> Dict[str, Any]:
    qn = norm(payload.question)
    sales = get_frame(cache, "sales")
    sales_idx = get_index(cache, "sales")
    month_key = extract_target_month(qn, sales_idx.months, payload.filters)
    stores = extract_store_filters(qn, payload.filters, sales_idx.stores)
    current_month = month_key
    base_month = same_month_ly(month_key) if has_any(qn, "ly", "año pasado", "ano pasado") else previous_month(month_key, sales_idx.months)
    cur_sales = apply_sales_domain_filter(apply_common_filters(sales, stores, current_month, payload.filters), "sales")
    base_sales = apply_sales_domain_filter(apply_common_filters(sales, stores, base_month, payload.filters), "sales") if base_month else sales.iloc[0:0]
    cur_delivery = apply_sales_domain_filter(apply_common_filters(sales, stores, current_month, payload.filters), "delivery")
    base_delivery = apply_sales_domain_filter(apply_common_filters(sales, stores, base_month, payload.filters), "delivery") if base_month else sales.iloc[0:0]
    sbf = get_frame(cache, "sbf")
    cur_sbf = apply_common_filters(sbf, stores, current_month, payload.filters)
    base_sbf = apply_common_filters(sbf, stores, base_month, payload.filters) if base_month else sbf.iloc[0:0]
    merma = get_frame(cache, "merma")
    cur_merma = apply_common_filters(merma, stores, current_month, payload.filters)

    sales_delta = pct_delta(float(cur_sales["Sales"].sum()), float(base_sales["Sales"].sum()) if not base_sales.empty else 0)
    del_delta = pct_delta(float(cur_delivery["Sales"].sum()), float(base_delivery["Sales"].sum()) if not base_delivery.empty else 0)
    sbf_delta = pct_delta(float(cur_sbf["TX"].sum()), float(base_sbf["TX"].sum()) if not base_sbf.empty else 0)
    merma_top = aggregate_group(cur_merma, "Description", "Merma").sort_values("Merma", ascending=False).head(3) if not cur_merma.empty else pd.DataFrame()

    summary_parts = [
        f"Ventas {month_label(current_month)}: {fmt_money(float(cur_sales['Sales'].sum()))} vs base {fmt_money(float(base_sales['Sales'].sum()) if not base_sales.empty else 0)} ⇒ {fmt_pct(sales_delta)}.",
        f"Delivery {month_label(current_month)}: {fmt_money(float(cur_delivery['Sales'].sum()))} vs base {fmt_money(float(base_delivery['Sales'].sum()) if not base_delivery.empty else 0)} ⇒ {fmt_pct(del_delta)}.",
        f"SBF {month_label(current_month)}: {fmt_num(float(cur_sbf['TX'].sum()))} TX vs base {fmt_num(float(base_sbf['TX'].sum()) if not base_sbf.empty else 0)} ⇒ {fmt_pct(sbf_delta)}.",
        f"Merma visible {month_label(current_month)}: {fmt_money(float(cur_merma['Merma'].sum()) if not cur_merma.empty else 0)}." + (f" Drivers: {summarise_top(merma_top, 'Description', 'Merma', limit=3)}." if not merma_top.empty else ""),
    ]
    return {
        "title": "Hallazgos ejecutivos",
        "summary": " ".join(summary_parts),
        "actions": default_actions("hallazgos"),
        "notes": build_notes(cache, current_month, stores, [f"Base comparativa: {month_label(base_month) if base_month else 'sin base'}"]),
        "chart": None,
    }


def resolve_query(cache: SourceCache, payload: QueryRequest) -> Dict[str, Any]:
    question = txt(payload.question)
    if not question:
        raise HTTPException(status_code=400, detail="La pregunta viene vacía.")
    qn = norm(question)
    active_tab = norm(payload.page_context.get("tab"))
    domain, conflict = pick_domain(qn, active_tab)
    if conflict:
        return {
            "title": "Consulta ambigua",
            "summary": f"La pregunta mezcla dos dominios distintos: {DOMAIN_LABELS.get(conflict[0], conflict[0])} y {DOMAIN_LABELS.get(conflict[1], conflict[1])}. Para responder bien, sepárala en dos consultas.",
            "actions": [
                f"Ejemplo 1: dame {DOMAIN_LABELS.get(conflict[0], conflict[0]).lower()} por tienda con delta vs LY.",
                f"Ejemplo 2: dame {DOMAIN_LABELS.get(conflict[1], conflict[1]).lower()} por tienda con delta vs LY.",
                "No mezcles dos motores de negocio en una misma pregunta si quieres cifras confiables.",
            ],
            "notes": [f"Fuente de verdad del chat: {cache.source}"],
            "chart": None,
        }
    if domain == "hallazgos" or has_any(qn, "hallazgo", "hallazgos", "lectura ejecutiva", "riesgo", "oportunidad"):
        return resolve_hallazgos(cache, payload)
    if domain in {"sales", "delivery", "innovation"}:
        return resolve_sales_like(cache, domain, payload)
    if domain == "sbf":
        return resolve_sbf(cache, payload)
    if domain in {"inventory", "merma", "cxc", "prd0"}:
        return resolve_inventory_merma_cxc_prd0(cache, domain, payload)
    return resolve_sales_like(cache, "sales", payload)


@app.get("/health")
def health() -> Dict[str, Any]:
    return {
        "ok": True,
        "version": APP_VERSION,
        "cache_ttl_seconds": CACHE_TTL_SECONDS,
        "default_data_url": DEFAULT_DATA_URL or None,
        "allowed_origins": allow_origins,
    }


@app.post("/api/query")
def query(payload: QueryRequest) -> Dict[str, Any]:
    source = txt(payload.data_url) or DEFAULT_DATA_URL
    if not source:
        raise HTTPException(status_code=400, detail="No llegó data_url y WORKBOOK_PUBLIC_URL no está configurado.")
    try:
        cache = get_source_cache(source)
        result = resolve_query(cache, payload)
        return {"ok": True, "engine": APP_VERSION, "source": source, "result": result}
    except HTTPException:
        raise
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"No pude descargar la fuente publicada: {exc}") from exc
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=f"No encontré el workbook: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error resolviendo la consulta: {exc}") from exc
